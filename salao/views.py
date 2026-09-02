import calendar
import csv
import io
import re
import unicodedata
import uuid
from collections import Counter, defaultdict
from itertools import combinations
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_DOWN, ROUND_HALF_UP

from openpyxl import Workbook
from openpyxl.styles import Font
from django.contrib import messages
from django.contrib.auth.decorators import user_passes_test
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import CharField, Count, F, Sum, Value
from django.db.models.deletion import ProtectedError
from django.db.models.functions import Coalesce
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.urls import reverse
from django.utils import timezone

from .models import (
    CompraEstoqueItemSalao,
    CompraEstoqueSalao,
    CategoriaDespesaSalao,
    ComissaoMensalSalao,
    DespesaSalao,
    FormaPagamentoSalao,
    LancamentoSalao,
    MovimentoEstoqueSalao,
    ProdutoSalao,
    ServicoSalao,
    SubcategoriaDespesaSalao,
    TaxaFormaPagamentoSalao,
    TransacaoIgnoradaSalao,
)


MONTH_OPTIONS = [(m, f"{m:02d}") for m in range(1, 13)]
_CODIGO_NUMERIC_CHUNKS_RE = re.compile(r'(\d+)')


def _salao_superuser_required(view_func):
    return user_passes_test(
        lambda user: user.is_authenticated and user.is_superuser,
        login_url='/admin/login/',
    )(view_func)


def _normalize_codigo(value: str) -> str:
    return (value or '').strip().upper()


def _codigo_natural_sort_key(value: str):
    normalized = (value or '').strip().upper()
    parts = _CODIGO_NUMERIC_CHUNKS_RE.split(normalized)
    return [(0, int(part)) if part.isdigit() else (1, part) for part in parts]


def _sort_produtos_por_codigo_natural(produtos):
    return sorted(produtos, key=lambda produto: _codigo_natural_sort_key(produto.codigo))


def _parse_competencia(request):
    today = date.today()
    raw_ano = request.POST.get('ano') or request.GET.get('ano')
    raw_mes = request.POST.get('mes') or request.GET.get('mes')

    try:
        ano = int(raw_ano) if raw_ano else today.year
    except (TypeError, ValueError):
        ano = today.year

    try:
        mes = int(raw_mes) if raw_mes else today.month
    except (TypeError, ValueError):
        mes = today.month

    if mes < 1 or mes > 12:
        mes = today.month

    if ano < 2000 or ano > 2100:
        ano = today.year

    return ano, mes


def _parse_day(request, ano, mes, clamp_on_overflow=False):
    raw_dia = request.POST.get('dia') or request.GET.get('dia')
    if raw_dia is None or raw_dia == '':
        return min(date.today().day, calendar.monthrange(ano, mes)[1])

    try:
        dia = int(raw_dia)
    except (TypeError, ValueError):
        return None

    ultimo_dia = calendar.monthrange(ano, mes)[1]
    if clamp_on_overflow and dia > ultimo_dia:
        return ultimo_dia

    if clamp_on_overflow and dia < 1:
        return 1

    if dia < 1 or dia > ultimo_dia:
        return None

    return dia


def _parse_decimal(raw_value, quantize_pattern='0.01'):
    normalized = (raw_value or '').strip().replace(',', '.')
    if not normalized:
        return None
    try:
        value = Decimal(normalized)
    except (InvalidOperation, ValueError):
        return None
    return value.quantize(Decimal(quantize_pattern), rounding=ROUND_HALF_UP)


def _parse_int_in_range(raw_value, default, min_value, max_value):
    try:
        value = int(raw_value)
    except (TypeError, ValueError):
        return default
    if value < min_value or value > max_value:
        return default
    return value


def _parse_checkbox(raw_value):
    return raw_value in ('on', '1', 'true', 'True')


def _parse_subcategoria_despesa(categoria_id, subcategoria_id):
    if not subcategoria_id:
        return None
    return SubcategoriaDespesaSalao.objects.filter(
        id=subcategoria_id,
        categoria_id=categoria_id,
        ativo=True,
    ).first()


def _redirect_lancamentos(ano, mes, dia):
    return redirect(f"{reverse('salao:lancamentos')}?ano={ano}&mes={mes}&dia={dia}")


def _redirect_despesas(ano, mes):
    return redirect(f"{reverse('salao:despesas')}?ano={ano}&mes={mes}")


def _redirect_dashboard(ano, mes):
    return redirect(f"{reverse('salao:dashboard')}?ano={ano}&mes={mes}")


def _redirect_conferencia(ano, mes, dia=None, retorno=''):
    """Volta para a conferência, na tela em que o usuário estava.

    `retorno` é a query string de origem. Conferir um dia inteiro ou marcar uma
    transação como fora do salão não pode ligar um filtro de dia que o usuário
    não pediu — mas se ele já estava filtrando o dia 1, continua filtrado.
    """
    if retorno:
        return redirect(f"{reverse('salao:conferencia')}?{retorno}")
    url = f"{reverse('salao:conferencia')}?ano={ano}&mes={mes}"
    if dia:
        url += f'&dia={dia}'
    return redirect(url)


def _redirect_grid_lancamentos(ano, mes):
    return redirect(f"{reverse('salao:grid_lancamentos')}?ano={ano}&mes={mes}")


def _redirect_grid_despesas(ano, mes):
    return redirect(f"{reverse('salao:grid_despesas')}?ano={ano}&mes={mes}")


def _redirect_servicos():
    return redirect(reverse('salao:servicos'))


def _redirect_categorias():
    return redirect(reverse('salao:categorias'))


def _redirect_pagamentos(forma_taxa_id=None):
    if forma_taxa_id:
        return redirect(f"{reverse('salao:pagamentos')}?forma_taxa={forma_taxa_id}")
    return redirect(reverse('salao:pagamentos'))


def _redirect_produtos():
    return redirect(reverse('salao:produtos'))


def _redirect_estoque(ano, mes):
    return redirect(f"{reverse('salao:estoque')}?ano={ano}&mes={mes}")


def _date_range_for_month(ano, mes):
    start = date(ano, mes, 1)
    end = date(ano, mes, calendar.monthrange(ano, mes)[1])
    return start, end


def _iter_months_backwards(ano, mes, quantidade):
    meses = []
    ano_cursor = ano
    mes_cursor = mes
    for _ in range(quantidade):
        meses.append((ano_cursor, mes_cursor))
        mes_cursor -= 1
        if mes_cursor == 0:
            mes_cursor = 12
            ano_cursor -= 1
    meses.reverse()
    return meses


def _build_year_options():
    current = date.today().year
    return [current - 1, current, current + 1, current + 2]


def _add_months_preserving_day(base_date, months_to_add):
    base_month_index = (base_date.month - 1) + months_to_add
    year = base_date.year + (base_month_index // 12)
    month = (base_month_index % 12) + 1
    day = min(base_date.day, calendar.monthrange(year, month)[1])
    return date(year, month, day)


def _split_amount_evenly(total, parcelas):
    parcela_base = (total / Decimal(parcelas)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    valores = [parcela_base for _ in range(parcelas)]
    diferenca = total - sum(valores)
    if diferenca != Decimal('0.00'):
        valores[-1] = (valores[-1] + diferenca).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    return valores


def _parse_parcelas(raw_value, default=1):
    return _parse_int_in_range(raw_value, default=default, min_value=1, max_value=12)


def _build_formas_catalogo(formas):
    taxas_qs = TaxaFormaPagamentoSalao.objects.select_related('forma_pagamento').filter(
        forma_pagamento__in=formas
    )
    taxas_por_forma = {}
    for taxa in taxas_qs:
        taxas_por_forma.setdefault(taxa.forma_pagamento_id, {})[str(taxa.parcelas)] = str(taxa.percentual)

    catalogo = []
    for forma in formas:
        catalogo.append(
            {
                'id': forma.id,
                'codigo': forma.codigo,
                'nome': forma.nome,
                'aceita_parcelamento': forma.aceita_parcelamento,
                'taxas': taxas_por_forma.get(forma.id, {}),
            }
        )
    return catalogo


def _forma_pagamento_nao_informado_ativa():
    return FormaPagamentoSalao.objects.filter(codigo='0', ativo=True).first()


def _calcular_liquido_com_taxa(valor_bruto, percentual):
    valor_taxa = (valor_bruto * percentual / Decimal('100.00')).quantize(
        Decimal('0.01'),
        rounding=ROUND_HALF_UP,
    )
    valor_liquido = (valor_bruto - valor_taxa).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    if valor_liquido < Decimal('0.00'):
        valor_liquido = Decimal('0.00')
    return valor_taxa, valor_liquido


def _calcular_bruto_com_taxa_repassada(valor_liquido_desejado, percentual):
    """Quanto cobrar do cliente para o salão receber `valor_liquido_desejado`.

    O caminho normal é o contrário: informa-se o bruto e desconta-se a taxa.
    Quando a taxa é repassada à cliente, quem manda é o líquido — cobra-se
    R$ 103,25 para receber R$ 100,00 limpos.

    Devolve o menor bruto cujo líquido alcança o alvo, porque arredondar o
    gross-up para baixo deixaria o salão um centavo abaixo do combinado.
    """
    if percentual <= Decimal('0.00'):
        return valor_liquido_desejado

    fator = Decimal('1.00') - (percentual / Decimal('100.00'))
    if fator <= Decimal('0.00'):
        return valor_liquido_desejado

    bruto = (valor_liquido_desejado / fator).quantize(Decimal('0.01'), rounding=ROUND_DOWN)
    while _calcular_liquido_com_taxa(bruto, percentual)[1] < valor_liquido_desejado:
        bruto += Decimal('0.01')
    return bruto


def _quantize_money(value):
    return value.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)


def _quantize_quantity(value):
    return value.quantize(Decimal('0.001'), rounding=ROUND_HALF_UP)


def _parse_compra_itens(request):
    produto_ids = request.POST.getlist('produto_id[]')
    quantidades = request.POST.getlist('quantidade[]')
    custos_unitarios = request.POST.getlist('custo_unitario[]')

    total_rows = max(len(produto_ids), len(quantidades), len(custos_unitarios))
    itens = []

    for idx in range(total_rows):
        raw_produto = (produto_ids[idx] if idx < len(produto_ids) else '').strip()
        raw_qtd = (quantidades[idx] if idx < len(quantidades) else '').strip()
        raw_custo = (custos_unitarios[idx] if idx < len(custos_unitarios) else '').strip()

        if not raw_produto and not raw_qtd and not raw_custo:
            continue

        if not raw_produto or not raw_qtd or not raw_custo:
            return None, f'Preencha produto, quantidade e custo unitário na linha {idx + 1}.'

        produto = ProdutoSalao.objects.filter(id=raw_produto, ativo=True).first()
        if not produto:
            return None, f'Produto inválido na linha {idx + 1}.'

        quantidade = _parse_decimal(raw_qtd, quantize_pattern='0.001')
        custo_unitario = _parse_decimal(raw_custo, quantize_pattern='0.01')

        if quantidade is None or quantidade <= Decimal('0.000'):
            return None, f'Quantidade inválida na linha {idx + 1}.'
        if custo_unitario is None or custo_unitario < Decimal('0.00'):
            return None, f'Custo unitário inválido na linha {idx + 1}.'

        custo_total = _quantize_money(quantidade * custo_unitario)
        itens.append(
            {
                'produto': produto,
                'quantidade': quantidade,
                'custo_unitario': custo_unitario,
                'custo_total': custo_total,
            }
        )

    if not itens:
        return None, 'Adicione ao menos um produto na compra com estoque.'

    return itens, None


def _rebuild_produto_from_movimentos(produto_id):
    produto = ProdutoSalao.objects.select_for_update().get(id=produto_id)
    movimentos = MovimentoEstoqueSalao.objects.filter(produto_id=produto.id).order_by('data', 'id')

    saldo = Decimal('0.000')
    custo_medio = Decimal('0.00')

    for movimento in movimentos:
        quantidade = movimento.quantidade or Decimal('0.000')
        if quantidade <= Decimal('0.000'):
            continue

        if movimento.tipo == MovimentoEstoqueSalao.TIPO_ENTRADA:
            custo_total_entrada = movimento.valor_custo_total or Decimal('0.00')
            valor_estoque_atual = _quantize_money(saldo * custo_medio)
            novo_saldo = saldo + quantidade
            if novo_saldo > Decimal('0.000'):
                custo_medio = _quantize_money((valor_estoque_atual + custo_total_entrada) / novo_saldo)
            saldo = novo_saldo
        elif movimento.tipo == MovimentoEstoqueSalao.TIPO_SAIDA:
            saldo = saldo - quantidade
            if saldo < Decimal('0.000'):
                saldo = Decimal('0.000')

    produto.saldo_atual = _quantize_quantity(saldo)
    produto.custo_medio_atual = _quantize_money(custo_medio if saldo > Decimal('0.000') else Decimal('0.00'))
    produto.save(update_fields=['saldo_atual', 'custo_medio_atual', 'atualizado_em'])


def _registrar_entrada_compra(compra, itens):
    produtos_para_rebuild = set()
    movimentos = []
    item_rows = []

    for item in itens:
        produto = ProdutoSalao.objects.select_for_update().get(id=item['produto'].id)
        quantidade = item['quantidade']
        custo_unitario = item['custo_unitario']
        custo_total = item['custo_total']

        valor_estoque_atual = _quantize_money(produto.saldo_atual * produto.custo_medio_atual)
        novo_saldo = produto.saldo_atual + quantidade
        novo_custo_medio = (
            _quantize_money((valor_estoque_atual + custo_total) / novo_saldo)
            if novo_saldo > Decimal('0.000')
            else Decimal('0.00')
        )

        produto.saldo_atual = _quantize_quantity(novo_saldo)
        produto.custo_medio_atual = novo_custo_medio
        produto.save(update_fields=['saldo_atual', 'custo_medio_atual', 'atualizado_em'])

        item_rows.append(
            CompraEstoqueItemSalao(
                compra=compra,
                produto=produto,
                quantidade=quantidade,
                custo_unitario=custo_unitario,
                custo_total=custo_total,
            )
        )

        movimentos.append(
            MovimentoEstoqueSalao(
                data=compra.data,
                produto=produto,
                tipo=MovimentoEstoqueSalao.TIPO_ENTRADA,
                motivo=MovimentoEstoqueSalao.MOTIVO_COMPRA,
                quantidade=quantidade,
                custo_unitario_aplicado=custo_unitario,
                valor_custo_total=custo_total,
                compra_estoque=compra,
                observacao=compra.observacao,
            )
        )
        produtos_para_rebuild.add(produto.id)

    CompraEstoqueItemSalao.objects.bulk_create(item_rows)
    MovimentoEstoqueSalao.objects.bulk_create(movimentos)
    for produto_id in produtos_para_rebuild:
        _rebuild_produto_from_movimentos(produto_id)


def _reverter_compra_estoque(compra):
    if not compra:
        return
    produto_ids = list(
        MovimentoEstoqueSalao.objects.filter(compra_estoque=compra).values_list('produto_id', flat=True).distinct()
    )
    MovimentoEstoqueSalao.objects.filter(compra_estoque=compra).delete()
    compra.delete()
    for produto_id in produto_ids:
        _rebuild_produto_from_movimentos(produto_id)


# ---------------------------------------------------------------------------
# Conferência bancária (CSV da adquirente x lançamentos)
# ---------------------------------------------------------------------------

CONFERENCIA_SESSION_KEY = 'conferencia_extrato'

_CSV_COLUNAS = {
    'data': 'Data e hora',
    'meio': 'Meio - Meio',
    'bandeira': 'Meio - Bandeira',
    'parcelas': 'Meio - Parcelas',
    'origem': 'Tipo - Origem',
    'identificador': 'Identificador',
    'status': 'Status',
    'bruto': 'Valor (R$)',
    'liquido': 'Líquido (R$)',
    'taxa_percentual': 'Taxa Aplicada - Aplicada(%)',
    'nome': 'Origem - Nome',
}


def _normalizar_texto(value):
    texto = unicodedata.normalize('NFKD', (value or '').strip().lower())
    return ''.join(ch for ch in texto if not unicodedata.combining(ch))


def _parse_valor_extrato(raw_value):
    """Aceita tanto '8.000,00' (pt-BR) quanto '3.14' (ponto decimal)."""
    texto = (raw_value or '').strip().replace("'", '').replace('R$', '')
    texto = texto.replace('\xa0', ' ').replace(' ', '').replace('-', '')
    if not texto:
        return Decimal('0.00')
    if ',' in texto:
        texto = texto.replace('.', '').replace(',', '.')
    try:
        return Decimal(texto).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    except InvalidOperation:
        return Decimal('0.00')


def _parse_parcelas_extrato(raw_value):
    somente_digitos = re.sub(r'\D', '', raw_value or '')
    return int(somente_digitos) if somente_digitos else 1


def _parse_extrato_csv(arquivo):
    """Lê o CSV da adquirente e devolve todas as transações do arquivo.

    Não filtra por competência: cada linha já carrega a própria data, então a
    tela decide o que mostrar. Assim o extrato importa certo independente do
    mês que estiver selecionado no filtro.
    """
    conteudo = arquivo.read()
    if isinstance(conteudo, bytes):
        conteudo = conteudo.decode('utf-8-sig', errors='replace')

    amostra = conteudo[:4096]
    try:
        dialeto = csv.Sniffer().sniff(amostra, delimiters=',;')
        delimitador = dialeto.delimiter
    except csv.Error:
        delimitador = ','

    leitor = csv.DictReader(io.StringIO(conteudo), delimiter=delimitador)
    if not leitor.fieldnames or _CSV_COLUNAS['data'] not in leitor.fieldnames:
        raise ValueError(
            'CSV fora do formato esperado. A primeira linha precisa ter as colunas do '
            'relatório da InfinitePay (Data e hora, Valor (R$), Status...).'
        )

    transacoes = []
    for linha in leitor:
        bruto_data = (linha.get(_CSV_COLUNAS['data']) or '').strip()
        if not bruto_data:
            continue
        try:
            data_hora = datetime.strptime(bruto_data[:16], '%d/%m/%Y %H:%M')
        except ValueError:
            try:
                data_hora = datetime.strptime(bruto_data[:10], '%d/%m/%Y')
            except ValueError:
                continue

        transacoes.append({
            'ano': data_hora.year,
            'mes': data_hora.month,
            'dia': data_hora.day,
            'hora': data_hora.strftime('%H:%M'),
            'meio': (linha.get(_CSV_COLUNAS['meio']) or '').strip(),
            'bandeira': (linha.get(_CSV_COLUNAS['bandeira']) or '').strip(),
            'parcelas': _parse_parcelas_extrato(linha.get(_CSV_COLUNAS['parcelas'])),
            'origem': (linha.get(_CSV_COLUNAS['origem']) or '').strip(),
            'identificador': (linha.get(_CSV_COLUNAS['identificador']) or '').strip(),
            'status': (linha.get(_CSV_COLUNAS['status']) or '').strip(),
            'bruto': str(_parse_valor_extrato(linha.get(_CSV_COLUNAS['bruto']))),
            'liquido': str(_parse_valor_extrato(linha.get(_CSV_COLUNAS['liquido']))),
            'taxa_percentual': str(_parse_valor_extrato(linha.get(_CSV_COLUNAS['taxa_percentual']))),
            'nome': (linha.get(_CSV_COLUNAS['nome']) or '').strip(),
        })

    return transacoes


# A taxa real varia por bandeira (elo sai ~1,8 p.p. acima de visa/mastercard),
# então só vale avisar quando a diferença para o cadastro é grande o bastante.
TOLERANCIA_TAXA_PP = Decimal('0.50')

# Quantos lançamentos podem ser somados para fechar uma única transação
# (cliente que paga 290 numa passada só, cobrindo um serviço de 170 e outro de 120).
MAX_LANCAMENTOS_POR_TRANSACAO = 3


class ItemConferencia:
    """Um serviço ou uma venda de produto, com a mesma cara para o pareamento.

    Os dois entram no mesmo extrato da adquirente, então a conferência precisa
    olhar para os dois. Guardam o valor em campos diferentes
    (`valor_bruto` x `valor_bruto_venda`), e é só isso que esta classe resolve.
    """

    SERVICO = 'servico'
    PRODUTO = 'produto'

    def __init__(self, origem, tipo):
        self.origem = origem
        self.tipo = tipo
        self.id = origem.id
        self.chave = f'{tipo}:{origem.id}'
        self.data = origem.data
        self.forma_pagamento = origem.forma_pagamento
        self.parcelas = origem.parcelas
        self.conferido = origem.conferido
        self.taxa_percentual_aplicada = origem.taxa_percentual_aplicada
        if tipo == self.SERVICO:
            self.valor_bruto = origem.valor_bruto
            self.descricao = f'{origem.servico.codigo} - {origem.servico.nome}'
            self.permuta = origem.permuta
            self.sem_comissao = origem.sem_comissao
            self.ajustavel = True
        else:
            self.valor_bruto = origem.valor_bruto_venda
            quantidade = origem.quantidade.normalize()
            self.descricao = f'{origem.produto.codigo} - {origem.produto.nome} ({quantidade}x)'
            self.permuta = False
            self.sem_comissao = False
            self.ajustavel = False


def _itens_conferencia_do_mes(ano, mes):
    """Serviços e vendas de produto do mês, na ordem em que aconteceram."""
    servicos = [
        ItemConferencia(lancamento, ItemConferencia.SERVICO)
        for lancamento in LancamentoSalao.objects.filter(data__year=ano, data__month=mes)
        .select_related('servico', 'forma_pagamento')
    ]
    produtos = [
        ItemConferencia(movimento, ItemConferencia.PRODUTO)
        for movimento in MovimentoEstoqueSalao.objects.filter(
            data__year=ano,
            data__month=mes,
            tipo=MovimentoEstoqueSalao.TIPO_SAIDA,
            motivo=MovimentoEstoqueSalao.MOTIVO_VENDA,
        ).select_related('produto', 'forma_pagamento')
    ]
    return sorted(servicos + produtos, key=lambda item: (item.data, item.tipo, item.id))


def _marcar_conferido(chave, conferido):
    """Marca um item da conferência, seja ele serviço ou venda de produto."""
    tipo, _, bruto_id = (chave or '').partition(':')
    modelo = {
        ItemConferencia.SERVICO: LancamentoSalao,
        ItemConferencia.PRODUTO: MovimentoEstoqueSalao,
    }.get(tipo)
    if modelo is None or not bruto_id.isdigit():
        return None
    origem = get_object_or_404(modelo, pk=int(bruto_id))
    origem.conferido = conferido
    origem.conferido_em = timezone.now() if conferido else None
    origem.save(update_fields=['conferido', 'conferido_em'])
    return origem


def _meio_do_lancamento(lancamento):
    return _normalizar_texto(
        lancamento.forma_pagamento.nome if lancamento.forma_pagamento else ''
    )


def _montar_par(lancamentos, transacao, criterio, diferenca):
    """Monta a linha de conferência de uma transação já casada.

    A comparação de taxa só faz sentido quando um único lançamento cobre a
    transação; somando vários, cada um pode ter a sua.
    """
    diferenca_taxa = None
    if len(lancamentos) == 1:
        diferenca_taxa = (
            Decimal(transacao['taxa_percentual']) - lancamentos[0].taxa_percentual_aplicada
        ).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    return {
        'lancamentos': lancamentos,
        'transacao': transacao,
        'criterio': criterio,
        'combinado': len(lancamentos) > 1,
        'entre_dias': False,
        'total_lancado': sum((l.valor_bruto for l in lancamentos), Decimal('0.00')),
        'todos_conferidos': all(l.conferido for l in lancamentos),
        'diferenca_valor': diferenca.quantize(Decimal('0.01')),
        'diferenca_taxa': diferenca_taxa,
        'taxa_divergente': diferenca_taxa is not None and abs(diferenca_taxa) > TOLERANCIA_TAXA_PP,
    }


def _parear_dia(lancamentos, transacoes, tolerancia=Decimal('1.00')):
    """Casa lançamentos do sistema com transações do extrato, dentro de um dia.

    O valor lançado no sistema ora é o que o cliente pagou (= bruto do extrato),
    ora é o valor do serviço com a taxa embutida no cartão (= líquido do
    extrato). Por isso cada passada testa os dois lados:

        1. um lançamento, valor exato (no bruto ou no líquido)
        2. um lançamento, dentro da tolerância
        3. vários lançamentos somados fechando uma transação só

    O meio de pagamento é sempre obrigatório: sem isso um Dinheiro de R$ 100,00
    casaria com um Crédito de R$ 99,91 só porque o valor ficou perto. Quando o
    meio não bate, o certo é sobrar dos dois lados e aparecer para conferência.
    """
    disponiveis = list(transacoes)
    pendentes = list(lancamentos)
    pares = []

    def avaliar(total, meio, limite):
        """Melhor transação disponível para um valor somado, ou None."""
        melhor = None
        for transacao in disponiveis:
            if _normalizar_texto(transacao['meio']) != meio:
                continue
            for criterio in ('bruto', 'liquido'):
                diferenca = total - Decimal(transacao[criterio])
                if abs(diferenca) <= limite and (melhor is None or abs(diferenca) < melhor[0]):
                    melhor = (abs(diferenca), transacao, criterio, diferenca)
        return melhor

    def registrar(grupo, escolha):
        _, transacao, criterio, diferenca = escolha
        disponiveis.remove(transacao)
        for lancamento in grupo:
            pendentes.remove(lancamento)
        pares.append(_montar_par(grupo, transacao, criterio, diferenca))

    for limite in (Decimal('0.00'), tolerancia):
        for lancamento in list(pendentes):
            escolha = avaliar(lancamento.valor_bruto, _meio_do_lancamento(lancamento), limite)
            if escolha:
                registrar([lancamento], escolha)

    # Sobras: tenta fechar uma transação somando lançamentos do mesmo meio.
    por_meio = defaultdict(list)
    for lancamento in pendentes:
        por_meio[_meio_do_lancamento(lancamento)].append(lancamento)

    for meio, grupo_meio in por_meio.items():
        for tamanho in range(2, MAX_LANCAMENTOS_POR_TRANSACAO + 1):
            while True:
                candidatos = [l for l in grupo_meio if l in pendentes]
                if len(candidatos) < tamanho:
                    break
                achou = None
                for combinacao in combinations(candidatos, tamanho):
                    total = sum((l.valor_bruto for l in combinacao), Decimal('0.00'))
                    escolha = avaliar(total, meio, tolerancia)
                    if escolha and (achou is None or escolha[0] < achou[1][0]):
                        achou = (list(combinacao), escolha)
                if achou is None:
                    break
                registrar(*achou)

    pares.sort(key=lambda par: par['lancamentos'][0].id)
    return pares, pendentes, disponiveis


def _parear_entre_dias(pendentes_por_dia, disponiveis_por_dia, tolerancia, janela_dias=7):
    """Fecha transações cujo pagamento não caiu no mesmo dia do atendimento.

    Um Pix de R$ 200,00 no dia 07 pode cobrir um atendimento do dia 01 e outro
    do dia 07. Roda só sobre o que sobrou do pareamento diário, exige o mesmo
    meio de pagamento e limita a distância entre as datas, porque cruzar dias
    aumenta muito a chance de casar duas coisas parecidas por acaso.

    Devolve os pares encontrados; as listas recebidas são esvaziadas do que foi
    consumido.
    """
    pares = []
    pendentes = [item for itens in pendentes_por_dia.values() for item in itens]

    for dia_transacao in sorted(disponiveis_por_dia):
        for transacao in list(disponiveis_por_dia[dia_transacao]):
            meio = _normalizar_texto(transacao['meio'])
            candidatos = [
                item for item in pendentes
                if _meio_do_lancamento(item) == meio
                and abs(item.data.day - dia_transacao) <= janela_dias
            ]
            achou = None
            for criterio in ('bruto', 'liquido'):
                alvo = Decimal(transacao[criterio])
                for tamanho in range(2, MAX_LANCAMENTOS_POR_TRANSACAO + 1):
                    if len(candidatos) < tamanho:
                        break
                    for combinacao in combinations(candidatos, tamanho):
                        if len({item.data for item in combinacao}) < 2:
                            continue  # mesmo dia já foi tentado no pareamento diário
                        diferenca = sum(
                            (item.valor_bruto for item in combinacao), Decimal('0.00')
                        ) - alvo
                        if abs(diferenca) <= tolerancia and (
                            achou is None or abs(diferenca) < abs(achou[2])
                        ):
                            achou = (list(combinacao), criterio, diferenca)
            if achou is None:
                continue

            combinacao, criterio, diferenca = achou
            disponiveis_por_dia[dia_transacao].remove(transacao)
            for item in combinacao:
                pendentes.remove(item)
                pendentes_por_dia[item.data.day].remove(item)
            par = _montar_par(combinacao, transacao, criterio, diferenca)
            par['entre_dias'] = True
            par['dia_transacao'] = dia_transacao
            pares.append(par)

    return pares


def _resumo_lancamentos_por_competencia(ano, mes, dia):
    inicio_mes, fim_mes = _date_range_for_month(ano, mes)
    data_fixa = date(ano, mes, dia)

    resumo_dia_qs = LancamentoSalao.objects.filter(data=data_fixa)
    resumo_mes_qs = LancamentoSalao.objects.filter(data__range=(inicio_mes, fim_mes))

    resumo_dia = {
        'qtd': resumo_dia_qs.count(),
        'total': resumo_dia_qs.aggregate(total=Sum('valor_cobrado'))['total'] or Decimal('0.00'),
    }
    resumo_mes = {
        'qtd': resumo_mes_qs.count(),
        'total': resumo_mes_qs.aggregate(total=Sum('valor_cobrado'))['total'] or Decimal('0.00'),
    }

    return resumo_dia, resumo_mes, inicio_mes, fim_mes, data_fixa


@_salao_superuser_required
def index(request):
    return redirect('salao:dashboard')


@_salao_superuser_required
def lancamentos(request):
    ano, mes = _parse_competencia(request)
    dia = _parse_day(request, ano, mes, clamp_on_overflow=(request.method == 'GET'))
    if dia is None:
        messages.error(request, 'Dia inválido para a competência selecionada.')
        dia = min(date.today().day, calendar.monthrange(ano, mes)[1])

    servicos_ativos = list(ServicoSalao.objects.filter(ativo=True).order_by('codigo'))
    formas_pagamento_ativas = list(FormaPagamentoSalao.objects.filter(ativo=True).order_by('codigo'))

    if request.GET.get('resumo') == '1' or request.GET.get('refresh') == '1':
        resumo_dia, resumo_mes, _, _, data_fixa = _resumo_lancamentos_por_competencia(ano, mes, dia)
        lancamentos_dia = (
            LancamentoSalao.objects.filter(data=data_fixa)
            .select_related('servico', 'forma_pagamento')
            .order_by('-id')[:30]
        )
        rows_html = render_to_string(
            'salao/partials/lancamentos_dia_rows.html',
            {
                'lancamentos_dia': lancamentos_dia,
                'ano': ano,
                'mes': mes,
                'dia': dia,
            },
            request=request,
        )
        return JsonResponse(
            {
                'ano': ano,
                'mes': mes,
                'dia': dia,
                'resumo_dia': {
                    'qtd': resumo_dia['qtd'],
                    'total': float(resumo_dia['total']),
                },
                'resumo_mes': {
                    'qtd': resumo_mes['qtd'],
                    'total': float(resumo_mes['total']),
                },
                'rows_html': rows_html,
            }
        )

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'create_lancamento':
            codigo = _normalize_codigo(request.POST.get('codigo'))
            codigo_forma = _normalize_codigo(request.POST.get('codigo_forma_pagamento'))
            permuta = _parse_checkbox(request.POST.get('permuta'))
            sem_comissao = not permuta and _parse_checkbox(request.POST.get('sem_comissao'))
            taxa_repassada = not permuta and _parse_checkbox(request.POST.get('taxa_repassada'))
            valor_bruto = _parse_decimal(request.POST.get('valor_bruto'))
            dia_post = _parse_day(request, ano, mes, clamp_on_overflow=False)

            if dia_post is None:
                messages.error(request, 'Dia inválido.')
                return _redirect_lancamentos(ano, mes, dia)

            if not codigo:
                messages.error(request, 'Informe o código do serviço.')
                return _redirect_lancamentos(ano, mes, dia_post)

            servico = ServicoSalao.objects.filter(codigo=codigo, ativo=True).first()
            if not servico:
                messages.error(request, f"Código '{codigo}' não encontrado entre os serviços ativos.")
                return _redirect_lancamentos(ano, mes, dia_post)

            if valor_bruto is None or valor_bruto < Decimal('0.00'):
                messages.error(request, 'Informe um valor bruto válido.')
                return _redirect_lancamentos(ano, mes, dia_post)

            if permuta:
                forma_pagamento = _forma_pagamento_nao_informado_ativa()
                if not forma_pagamento:
                    messages.error(
                        request,
                        'Forma de pagamento "0 - Não informado" não está ativa. Ajuste em Pagamentos.',
                    )
                    return _redirect_lancamentos(ano, mes, dia_post)
                parcelas = 1
                taxa_percentual = Decimal('0.00')
                valor_taxa = Decimal('0.00')
                valor_liquido = valor_bruto
            else:
                if not codigo_forma:
                    messages.error(request, 'Informe o código da forma de pagamento.')
                    return _redirect_lancamentos(ano, mes, dia_post)

                forma_pagamento = FormaPagamentoSalao.objects.filter(codigo=codigo_forma, ativo=True).first()
                if not forma_pagamento:
                    messages.error(request, f"Código de pagamento '{codigo_forma}' não encontrado.")
                    return _redirect_lancamentos(ano, mes, dia_post)

                parcelas = _parse_parcelas(request.POST.get('parcelas'), default=1)
                if not forma_pagamento.aceita_parcelamento:
                    parcelas = 1

                taxa = TaxaFormaPagamentoSalao.objects.filter(
                    forma_pagamento=forma_pagamento,
                    parcelas=parcelas,
                ).first()
                if not taxa:
                    messages.error(request, 'Taxa não cadastrada para essa forma de pagamento e parcela.')
                    return _redirect_lancamentos(ano, mes, dia_post)
                taxa_percentual = taxa.percentual
                if taxa_repassada:
                    # O valor digitado é o que o salão quer receber; a taxa vai por cima.
                    valor_bruto = _calcular_bruto_com_taxa_repassada(valor_bruto, taxa_percentual)
                valor_taxa, valor_liquido = _calcular_liquido_com_taxa(valor_bruto, taxa_percentual)

            LancamentoSalao.objects.create(
                data=date(ano, mes, dia_post),
                servico=servico,
                forma_pagamento=forma_pagamento,
                parcelas=parcelas,
                permuta=permuta,
                sem_comissao=sem_comissao,
                taxa_repassada=taxa_repassada,
                valor_bruto=valor_bruto,
                taxa_percentual_aplicada=taxa_percentual,
                valor_taxa=valor_taxa,
                valor_cobrado=valor_liquido,
            )
            if permuta:
                messages.success(request, f'Permuta salva. Valor integral: R$ {valor_liquido}.')
            elif sem_comissao:
                messages.success(request, f'Lançamento sem comissão salvo. Líquido: R$ {valor_liquido}.')
            elif taxa_repassada:
                messages.success(
                    request,
                    f'Lançamento salvo. Cobrar da cliente: R$ {valor_bruto} '
                    f'(taxa {taxa_percentual}%). Líquido: R$ {valor_liquido}.',
                )
            else:
                messages.success(request, f'Lançamento salvo. Líquido: R$ {valor_liquido}.')
            return _redirect_lancamentos(ano, mes, dia_post)

        if action == 'delete_lancamento':
            lancamento_id = request.POST.get('lancamento_id')
            lancamento = get_object_or_404(LancamentoSalao, id=lancamento_id)
            lancamento.delete()
            messages.success(request, 'Lançamento removido com sucesso.')
            return _redirect_lancamentos(ano, mes, dia)

        if action == 'update_lancamento':
            lancamento_id = request.POST.get('lancamento_id')
            lancamento = get_object_or_404(LancamentoSalao, id=lancamento_id)
            permuta = _parse_checkbox(request.POST.get('permuta'))
            sem_comissao = not permuta and _parse_checkbox(request.POST.get('sem_comissao'))
            taxa_repassada = not permuta and _parse_checkbox(request.POST.get('taxa_repassada'))

            raw_data = request.POST.get('data')
            try:
                ano_edit, mes_edit, dia_edit = [int(part) for part in raw_data.split('-')]
                data_editada = date(ano_edit, mes_edit, dia_edit)
            except (TypeError, ValueError, AttributeError):
                messages.error(request, 'Data inválida para edição do lançamento.')
                return _redirect_lancamentos(ano, mes, dia)

            servico_id = request.POST.get('servico_id')
            servico = ServicoSalao.objects.filter(id=servico_id, ativo=True).first()
            if not servico:
                messages.error(request, 'Selecione um serviço ativo válido para edição.')
                return _redirect_lancamentos(ano, mes, dia)

            valor_bruto = _parse_decimal(request.POST.get('valor_bruto'))
            if valor_bruto is None or valor_bruto < Decimal('0.00'):
                messages.error(request, 'Informe um valor bruto válido para edição.')
                return _redirect_lancamentos(ano, mes, dia)

            if permuta:
                forma_pagamento = _forma_pagamento_nao_informado_ativa()
                if not forma_pagamento:
                    messages.error(
                        request,
                        'Forma de pagamento "0 - Não informado" não está ativa. Ajuste em Pagamentos.',
                    )
                    return _redirect_lancamentos(ano, mes, dia)
                parcelas = 1
                taxa_percentual = Decimal('0.00')
                valor_taxa = Decimal('0.00')
                valor_liquido = valor_bruto
            else:
                forma_pagamento_id = request.POST.get('forma_pagamento_id')
                forma_pagamento = FormaPagamentoSalao.objects.filter(id=forma_pagamento_id, ativo=True).first()
                if not forma_pagamento:
                    messages.error(request, 'Selecione uma forma de pagamento ativa válida.')
                    return _redirect_lancamentos(ano, mes, dia)

                parcelas = _parse_parcelas(request.POST.get('parcelas'), default=1)
                if not forma_pagamento.aceita_parcelamento:
                    parcelas = 1

                taxa = TaxaFormaPagamentoSalao.objects.filter(
                    forma_pagamento=forma_pagamento,
                    parcelas=parcelas,
                ).first()
                if not taxa:
                    messages.error(request, 'Taxa não cadastrada para essa forma de pagamento e parcela.')
                    return _redirect_lancamentos(ano, mes, dia)
                taxa_percentual = taxa.percentual
                if taxa_repassada:
                    # O valor digitado é o que o salão quer receber; a taxa vai por cima.
                    valor_bruto = _calcular_bruto_com_taxa_repassada(valor_bruto, taxa_percentual)
                valor_taxa, valor_liquido = _calcular_liquido_com_taxa(valor_bruto, taxa_percentual)

            lancamento.data = data_editada
            lancamento.servico = servico
            lancamento.forma_pagamento = forma_pagamento
            lancamento.parcelas = parcelas
            lancamento.permuta = permuta
            lancamento.sem_comissao = sem_comissao
            lancamento.taxa_repassada = taxa_repassada
            lancamento.valor_bruto = valor_bruto
            lancamento.taxa_percentual_aplicada = taxa_percentual
            lancamento.valor_taxa = valor_taxa
            lancamento.valor_cobrado = valor_liquido
            lancamento.save(
                update_fields=[
                    'data',
                    'servico',
                    'forma_pagamento',
                    'parcelas',
                    'permuta',
                    'sem_comissao',
                    'taxa_repassada',
                    'valor_bruto',
                    'taxa_percentual_aplicada',
                    'valor_taxa',
                    'valor_cobrado',
                    'atualizado_em',
                ]
            )
            if permuta:
                messages.success(request, 'Permuta atualizada com sucesso.')
            elif sem_comissao:
                messages.success(request, 'Lançamento sem comissão atualizado com sucesso.')
            else:
                messages.success(request, 'Lançamento atualizado com sucesso.')
            return _redirect_lancamentos(ano, mes, dia)

    resumo_dia, resumo_mes, inicio_mes, fim_mes, data_fixa = _resumo_lancamentos_por_competencia(ano, mes, dia)

    lancamentos_mes = (
        LancamentoSalao.objects.filter(data__range=(inicio_mes, fim_mes))
        .select_related('servico')
        .order_by('-data', '-id')[:120]
    )

    lancamentos_dia = (
        LancamentoSalao.objects.filter(data=data_fixa)
        .select_related('servico', 'forma_pagamento')
        .order_by('-id')[:30]
    )

    edit_lancamento = None
    edit_lancamento_id = request.GET.get('edit')
    if edit_lancamento_id:
        edit_lancamento = LancamentoSalao.objects.filter(id=edit_lancamento_id).select_related(
            'servico', 'forma_pagamento'
        ).first()

    context = {
        'active_tab': 'lancamentos',
        'ano': ano,
        'mes': mes,
        'dia': dia,
        'month_options': MONTH_OPTIONS,
        'year_options': _build_year_options(),
        'servicos_ativos': servicos_ativos,
        'formas_pagamento_ativas': formas_pagamento_ativas,
        'servicos_catalogo': [
            {
                'id': servico.id,
                'codigo': servico.codigo,
                'nome': servico.nome,
                'valor_padrao': str(servico.valor_padrao),
            }
            for servico in servicos_ativos
        ],
        'formas_catalogo': _build_formas_catalogo(formas_pagamento_ativas),
        'lancamentos_mes': lancamentos_mes,
        'lancamentos_dia': lancamentos_dia,
        'edit_lancamento': edit_lancamento,
        'resumo_dia': resumo_dia,
        'resumo_mes': resumo_mes,
    }
    return render(request, 'salao/lancamentos.html', context)


@_salao_superuser_required
def despesas(request):
    ano, mes = _parse_competencia(request)
    categorias_ativas = list(
        CategoriaDespesaSalao.objects.filter(ativo=True).order_by('nome')
    )
    produtos_ativos = _sort_produtos_por_codigo_natural(
        list(ProdutoSalao.objects.filter(ativo=True))
    )

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'create_despesa':
            raw_data = request.POST.get('data')
            try:
                ano_d, mes_d, dia_d = [int(part) for part in raw_data.split('-')]
                data_despesa = date(ano_d, mes_d, dia_d)
            except (TypeError, ValueError, AttributeError):
                messages.error(request, 'Data inválida.')
                return _redirect_despesas(ano, mes)

            categoria_id = request.POST.get('categoria_id')
            categoria = CategoriaDespesaSalao.objects.filter(id=categoria_id, ativo=True).first()
            if not categoria:
                messages.error(request, 'Selecione uma categoria ativa.')
                return _redirect_despesas(ano, mes)
            subcategoria = _parse_subcategoria_despesa(categoria.id, request.POST.get('subcategoria_id'))
            if request.POST.get('subcategoria_id') and not subcategoria:
                messages.error(request, 'Selecione uma subcategoria ativa válida para a categoria.')
                return _redirect_despesas(ano, mes)

            observacao = (request.POST.get('observacao') or '').strip()
            parcelas = _parse_parcelas(request.POST.get('parcelas'), default=1)
            gera_estoque = _parse_checkbox(request.POST.get('gera_estoque'))

            itens = []
            custos_adicionais = Decimal('0.00')
            if gera_estoque:
                itens, erro_itens = _parse_compra_itens(request)
                if erro_itens:
                    messages.error(request, erro_itens)
                    return _redirect_despesas(ano, mes)
                raw_custos_adicionais = (request.POST.get('custos_adicionais') or '').strip()
                if raw_custos_adicionais:
                    custos_adicionais = _parse_decimal(raw_custos_adicionais)
                    if custos_adicionais is None:
                        messages.error(request, 'Informe um valor válido para custos adicionais.')
                        return _redirect_despesas(ano, mes)
                if custos_adicionais < Decimal('0.00'):
                    messages.error(request, 'Informe um valor válido para custos adicionais.')
                    return _redirect_despesas(ano, mes)
                valor_itens = _quantize_money(sum((item['custo_total'] for item in itens), Decimal('0.00')))
                valor = _quantize_money(valor_itens + custos_adicionais)
            else:
                valor = _parse_decimal(request.POST.get('valor'))
                if valor is None or valor < Decimal('0.00'):
                    messages.error(request, 'Informe um valor válido.')
                    return _redirect_despesas(ano, mes)

            grupo_parcelamento_id = uuid.uuid4() if parcelas > 1 else None
            valores_parcelados = _split_amount_evenly(valor, parcelas)

            with transaction.atomic():
                compra_estoque = None
                if gera_estoque:
                    compra_estoque = CompraEstoqueSalao.objects.create(
                        data=data_despesa,
                        categoria_fornecedor=categoria,
                        valor_total=valor,
                        custos_adicionais=custos_adicionais,
                        parcelas_total=parcelas,
                        grupo_parcelamento_id=grupo_parcelamento_id,
                        observacao=observacao,
                    )
                    _registrar_entrada_compra(compra_estoque, itens)

                despesas_para_criar = []
                for idx in range(parcelas):
                    data_parcela = _add_months_preserving_day(data_despesa, idx)
                    despesas_para_criar.append(
                        DespesaSalao(
                            data=data_parcela,
                            categoria=categoria,
                            subcategoria=subcategoria,
                            gera_estoque=gera_estoque,
                            compra_estoque=compra_estoque,
                            valor=valores_parcelados[idx],
                            observacao=observacao,
                            grupo_parcelamento_id=grupo_parcelamento_id,
                            parcela_numero=idx + 1,
                            parcelas_total=parcelas,
                        )
                    )
                DespesaSalao.objects.bulk_create(despesas_para_criar)

            if parcelas > 1:
                messages.success(request, f'Despesa parcelada salva em {parcelas}x.')
            else:
                messages.success(request, 'Despesa salva com sucesso.')
            return _redirect_despesas(ano, mes)

        if action == 'delete_despesa':
            despesa_id = request.POST.get('despesa_id')
            despesa = get_object_or_404(DespesaSalao, id=despesa_id)

            if despesa.gera_estoque and despesa.parcelas_total > 1 and despesa.grupo_parcelamento_id:
                messages.error(
                    request,
                    'Compra com estoque parcelada deve ser removida pelo botão "Excluir grupo".',
                )
                return _redirect_despesas(ano, mes)

            with transaction.atomic():
                if despesa.gera_estoque and despesa.compra_estoque:
                    outras_despesas = DespesaSalao.objects.filter(compra_estoque=despesa.compra_estoque).exclude(
                        id=despesa.id
                    )
                    if outras_despesas.exists():
                        messages.error(
                            request,
                            'Essa compra possui outras parcelas. Use "Excluir grupo" para remover tudo.',
                        )
                        return _redirect_despesas(ano, mes)
                    _reverter_compra_estoque(despesa.compra_estoque)

                despesa.delete()

            messages.success(request, 'Despesa removida com sucesso.')
            return _redirect_despesas(ano, mes)

        if action == 'delete_despesa_grupo':
            grupo_id = request.POST.get('grupo_parcelamento_id')
            try:
                grupo_uuid = uuid.UUID(str(grupo_id))
            except (TypeError, ValueError):
                messages.error(request, 'Grupo de parcelamento inválido.')
                return _redirect_despesas(ano, mes)

            with transaction.atomic():
                despesas_grupo = DespesaSalao.objects.filter(grupo_parcelamento_id=grupo_uuid)
                compras_ids = list(
                    despesas_grupo.filter(gera_estoque=True, compra_estoque__isnull=False)
                    .values_list('compra_estoque_id', flat=True)
                    .distinct()
                )
                for compra_id in compras_ids:
                    compra = CompraEstoqueSalao.objects.filter(id=compra_id).first()
                    if compra:
                        _reverter_compra_estoque(compra)
                deleted_count, _ = despesas_grupo.delete()

            if deleted_count > 0:
                messages.success(request, 'Grupo de despesas parceladas removido com sucesso.')
            else:
                messages.warning(request, 'Nenhuma despesa encontrada para o grupo informado.')
            return _redirect_despesas(ano, mes)

        if action == 'update_despesa':
            despesa_id = request.POST.get('despesa_id')
            despesa = get_object_or_404(DespesaSalao, id=despesa_id)
            if despesa.gera_estoque:
                messages.error(
                    request,
                    'Edição direta desativada para compras com estoque. Exclua e relance para manter o histórico.',
                )
                return _redirect_despesas(ano, mes)

            raw_data = request.POST.get('data')
            try:
                ano_d, mes_d, dia_d = [int(part) for part in raw_data.split('-')]
                data_editada = date(ano_d, mes_d, dia_d)
            except (TypeError, ValueError, AttributeError):
                messages.error(request, 'Data inválida para edição.')
                return _redirect_despesas(ano, mes)

            categoria_id = request.POST.get('categoria_id')
            categoria = CategoriaDespesaSalao.objects.filter(id=categoria_id, ativo=True).first()
            if not categoria:
                messages.error(request, 'Selecione uma categoria ativa válida.')
                return _redirect_despesas(ano, mes)
            subcategoria = _parse_subcategoria_despesa(categoria.id, request.POST.get('subcategoria_id'))
            if request.POST.get('subcategoria_id') and not subcategoria:
                messages.error(request, 'Selecione uma subcategoria ativa válida para a categoria.')
                return _redirect_despesas(ano, mes)

            valor = _parse_decimal(request.POST.get('valor'))
            if valor is None or valor < Decimal('0.00'):
                messages.error(request, 'Informe um valor válido para edição.')
                return _redirect_despesas(ano, mes)

            despesa.data = data_editada
            despesa.categoria = categoria
            despesa.subcategoria = subcategoria
            despesa.valor = valor
            despesa.observacao = (request.POST.get('observacao') or '').strip()
            despesa.save(
                update_fields=['data', 'categoria', 'subcategoria', 'valor', 'observacao', 'atualizado_em']
            )
            messages.success(request, 'Despesa atualizada com sucesso.')
            return _redirect_despesas(ano, mes)

    inicio_mes, fim_mes = _date_range_for_month(ano, mes)
    despesas_mes = (
        DespesaSalao.objects.filter(data__range=(inicio_mes, fim_mes))
        .select_related('categoria', 'subcategoria', 'compra_estoque')
        .order_by('data', 'parcela_numero', 'id')[:300]
    )

    edit_despesa = None
    edit_despesa_id = request.GET.get('edit')
    if edit_despesa_id:
        edit_despesa = DespesaSalao.objects.filter(id=edit_despesa_id).select_related(
            'categoria', 'subcategoria'
        ).first()

    subcategorias_ativas = list(
        SubcategoriaDespesaSalao.objects.filter(ativo=True)
        .select_related('categoria')
        .order_by('categoria__nome', 'nome')
    )

    context = {
        'active_tab': 'despesas',
        'ano': ano,
        'mes': mes,
        'month_options': MONTH_OPTIONS,
        'year_options': _build_year_options(),
        'categorias_ativas': categorias_ativas,
        'subcategorias_ativas': subcategorias_ativas,
        'subcategorias_catalogo': [
            {'id': sub.id, 'categoria_id': sub.categoria_id, 'nome': sub.nome}
            for sub in subcategorias_ativas
        ],
        'produtos_ativos': produtos_ativos,
        'despesas_mes': despesas_mes,
        'edit_despesa': edit_despesa,
        'data_padrao': date(ano, mes, min(date.today().day, calendar.monthrange(ano, mes)[1])),
    }
    return render(request, 'salao/despesas.html', context)


@_salao_superuser_required
def servicos(request):
    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'create_servico':
            codigo = _normalize_codigo(request.POST.get('codigo'))
            nome = (request.POST.get('nome') or '').strip()
            valor_padrao = _parse_decimal(request.POST.get('valor_padrao'))
            ativo = _parse_checkbox(request.POST.get('ativo'))

            if not codigo or not nome or valor_padrao is None:
                messages.error(request, 'Preencha código, nome e valor padrão válidos.')
                return _redirect_servicos()

            if ServicoSalao.objects.filter(codigo=codigo).exists():
                messages.error(request, f'Já existe serviço com código {codigo}.')
                return _redirect_servicos()

            ServicoSalao.objects.create(
                codigo=codigo,
                nome=nome,
                valor_padrao=valor_padrao,
                ativo=ativo,
            )
            messages.success(request, 'Serviço criado com sucesso.')
            return _redirect_servicos()

        if action == 'update_servico':
            servico_id = request.POST.get('servico_id')
            servico = get_object_or_404(ServicoSalao, id=servico_id)

            codigo = _normalize_codigo(request.POST.get('codigo'))
            nome = (request.POST.get('nome') or '').strip()
            valor_padrao = _parse_decimal(request.POST.get('valor_padrao'))
            ativo = _parse_checkbox(request.POST.get('ativo'))

            if not codigo or not nome or valor_padrao is None:
                messages.error(request, 'Preencha código, nome e valor padrão válidos.')
                return _redirect_servicos()

            if ServicoSalao.objects.exclude(id=servico.id).filter(codigo=codigo).exists():
                messages.error(request, f'Já existe outro serviço com código {codigo}.')
                return _redirect_servicos()

            servico.codigo = codigo
            servico.nome = nome
            servico.valor_padrao = valor_padrao
            servico.ativo = ativo
            servico.save()
            messages.success(request, 'Serviço atualizado com sucesso.')
            return _redirect_servicos()

        if action == 'delete_servico':
            servico_id = request.POST.get('servico_id')
            servico = get_object_or_404(ServicoSalao, id=servico_id)
            try:
                servico.delete()
                messages.success(request, 'Serviço removido com sucesso.')
            except ProtectedError:
                messages.error(
                    request,
                    'Não foi possível remover: esse serviço possui lançamentos vinculados.',
                )
            return _redirect_servicos()

    servicos_qs = ServicoSalao.objects.all().order_by('codigo')
    edit_id = request.GET.get('edit')
    edit_servico = ServicoSalao.objects.filter(id=edit_id).first() if edit_id else None

    context = {
        'active_tab': 'servicos',
        'servicos': servicos_qs,
        'edit_servico': edit_servico,
    }
    return render(request, 'salao/servicos.html', context)


@_salao_superuser_required
def categorias(request):
    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'create_categoria':
            nome = (request.POST.get('nome') or '').strip()
            ativo = _parse_checkbox(request.POST.get('ativo'))

            if not nome:
                messages.error(request, 'Informe o nome da categoria.')
                return _redirect_categorias()

            if CategoriaDespesaSalao.objects.filter(nome__iexact=nome).exists():
                messages.error(request, f'Categoria "{nome}" já existe.')
                return _redirect_categorias()

            CategoriaDespesaSalao.objects.create(
                nome=nome,
                ativo=ativo,
            )
            messages.success(request, 'Categoria criada com sucesso.')
            return _redirect_categorias()

        if action == 'update_categoria':
            categoria_id = request.POST.get('categoria_id')
            categoria = get_object_or_404(CategoriaDespesaSalao, id=categoria_id)

            nome = (request.POST.get('nome') or '').strip()
            ativo = _parse_checkbox(request.POST.get('ativo'))

            if not nome:
                messages.error(request, 'Informe o nome da categoria.')
                return _redirect_categorias()

            if CategoriaDespesaSalao.objects.exclude(id=categoria.id).filter(nome__iexact=nome).exists():
                messages.error(request, f'Categoria "{nome}" já existe.')
                return _redirect_categorias()

            categoria.nome = nome
            categoria.ativo = ativo
            categoria.save()
            messages.success(request, 'Categoria atualizada com sucesso.')
            return _redirect_categorias()

        if action == 'delete_categoria':
            categoria_id = request.POST.get('categoria_id')
            categoria = get_object_or_404(CategoriaDespesaSalao, id=categoria_id)
            try:
                categoria.delete()
                messages.success(request, 'Categoria removida com sucesso.')
            except ProtectedError:
                messages.error(
                    request,
                    'Não foi possível remover: essa categoria possui despesas vinculadas.',
                )
            return _redirect_categorias()

        if action == 'create_subcategoria':
            categoria_id = request.POST.get('categoria_id')
            categoria = CategoriaDespesaSalao.objects.filter(id=categoria_id).first()
            nome = (request.POST.get('nome') or '').strip()
            ativo = _parse_checkbox(request.POST.get('ativo'))
            if not categoria:
                messages.error(request, 'Selecione uma categoria válida para a subcategoria.')
                return _redirect_categorias()
            if not nome:
                messages.error(request, 'Informe o nome da subcategoria.')
                return _redirect_categorias()
            if SubcategoriaDespesaSalao.objects.filter(categoria=categoria, nome__iexact=nome).exists():
                messages.error(request, f'Subcategoria "{nome}" já existe nessa categoria.')
                return _redirect_categorias()
            SubcategoriaDespesaSalao.objects.create(
                categoria=categoria,
                nome=nome,
                ativo=ativo,
            )
            messages.success(request, 'Subcategoria criada com sucesso.')
            return _redirect_categorias()

        if action == 'update_subcategoria':
            subcategoria_id = request.POST.get('subcategoria_id')
            subcategoria = get_object_or_404(SubcategoriaDespesaSalao, id=subcategoria_id)
            categoria_id = request.POST.get('categoria_id')
            categoria = CategoriaDespesaSalao.objects.filter(id=categoria_id).first()
            nome = (request.POST.get('nome') or '').strip()
            ativo = _parse_checkbox(request.POST.get('ativo'))
            if not categoria:
                messages.error(request, 'Selecione uma categoria válida para a subcategoria.')
                return _redirect_categorias()
            if not nome:
                messages.error(request, 'Informe o nome da subcategoria.')
                return _redirect_categorias()
            if SubcategoriaDespesaSalao.objects.exclude(id=subcategoria.id).filter(
                categoria=categoria, nome__iexact=nome
            ).exists():
                messages.error(request, f'Subcategoria "{nome}" já existe nessa categoria.')
                return _redirect_categorias()
            subcategoria.categoria = categoria
            subcategoria.nome = nome
            subcategoria.ativo = ativo
            subcategoria.save(update_fields=['categoria', 'nome', 'ativo', 'atualizado_em'])
            messages.success(request, 'Subcategoria atualizada com sucesso.')
            return _redirect_categorias()

        if action == 'delete_subcategoria':
            subcategoria_id = request.POST.get('subcategoria_id')
            subcategoria = get_object_or_404(SubcategoriaDespesaSalao, id=subcategoria_id)
            try:
                subcategoria.delete()
                messages.success(request, 'Subcategoria removida com sucesso.')
            except ProtectedError:
                messages.error(
                    request,
                    'Não foi possível remover: essa subcategoria possui despesas vinculadas.',
                )
            return _redirect_categorias()

    categorias_qs = CategoriaDespesaSalao.objects.all().order_by('nome')
    edit_id = request.GET.get('edit')
    edit_categoria = CategoriaDespesaSalao.objects.filter(id=edit_id).first() if edit_id else None
    subcategorias_qs = SubcategoriaDespesaSalao.objects.select_related('categoria').order_by(
        'categoria__nome', 'nome'
    )
    edit_subcategoria_id = request.GET.get('edit_subcategoria')
    edit_subcategoria = (
        SubcategoriaDespesaSalao.objects.select_related('categoria').filter(id=edit_subcategoria_id).first()
        if edit_subcategoria_id
        else None
    )

    context = {
        'active_tab': 'categorias',
        'categorias': categorias_qs,
        'edit_categoria': edit_categoria,
        'subcategorias': subcategorias_qs,
        'edit_subcategoria': edit_subcategoria,
    }
    return render(request, 'salao/categorias.html', context)


@_salao_superuser_required
def produtos(request):
    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'create_produto':
            codigo = _normalize_codigo(request.POST.get('codigo'))
            nome = (request.POST.get('nome') or '').strip()
            unidade = _normalize_codigo(request.POST.get('unidade') or 'UN')
            valor_venda_padrao = _parse_decimal(request.POST.get('valor_venda_padrao'))
            estoque_minimo = _parse_decimal(request.POST.get('estoque_minimo'), quantize_pattern='0.001')
            ativo = _parse_checkbox(request.POST.get('ativo'))

            if not codigo or not nome:
                messages.error(request, 'Informe código e nome do produto.')
                return _redirect_produtos()
            if valor_venda_padrao is None or valor_venda_padrao < Decimal('0.00'):
                messages.error(request, 'Informe um valor de venda padrão válido.')
                return _redirect_produtos()
            if estoque_minimo is None or estoque_minimo < Decimal('0.000'):
                messages.error(request, 'Informe um estoque mínimo válido.')
                return _redirect_produtos()
            if ProdutoSalao.objects.filter(codigo=codigo).exists():
                messages.error(request, f'Já existe produto com código {codigo}.')
                return _redirect_produtos()

            ProdutoSalao.objects.create(
                codigo=codigo,
                nome=nome,
                unidade=unidade,
                valor_venda_padrao=valor_venda_padrao,
                estoque_minimo=estoque_minimo,
                ativo=ativo,
            )
            messages.success(request, 'Produto cadastrado com sucesso.')
            return _redirect_produtos()

        if action == 'update_produto':
            produto_id = request.POST.get('produto_id')
            produto = get_object_or_404(ProdutoSalao, id=produto_id)

            codigo = _normalize_codigo(request.POST.get('codigo'))
            nome = (request.POST.get('nome') or '').strip()
            unidade = _normalize_codigo(request.POST.get('unidade') or 'UN')
            valor_venda_padrao = _parse_decimal(request.POST.get('valor_venda_padrao'))
            estoque_minimo = _parse_decimal(request.POST.get('estoque_minimo'), quantize_pattern='0.001')
            ativo = _parse_checkbox(request.POST.get('ativo'))

            if not codigo or not nome:
                messages.error(request, 'Informe código e nome do produto.')
                return _redirect_produtos()
            if valor_venda_padrao is None or valor_venda_padrao < Decimal('0.00'):
                messages.error(request, 'Informe um valor de venda padrão válido.')
                return _redirect_produtos()
            if estoque_minimo is None or estoque_minimo < Decimal('0.000'):
                messages.error(request, 'Informe um estoque mínimo válido.')
                return _redirect_produtos()
            if ProdutoSalao.objects.exclude(id=produto.id).filter(codigo=codigo).exists():
                messages.error(request, f'Já existe outro produto com código {codigo}.')
                return _redirect_produtos()

            produto.codigo = codigo
            produto.nome = nome
            produto.unidade = unidade
            produto.valor_venda_padrao = valor_venda_padrao
            produto.estoque_minimo = estoque_minimo
            produto.ativo = ativo
            produto.save(
                update_fields=[
                    'codigo',
                    'nome',
                    'unidade',
                    'valor_venda_padrao',
                    'estoque_minimo',
                    'ativo',
                    'atualizado_em',
                ]
            )
            messages.success(request, 'Produto atualizado com sucesso.')
            return _redirect_produtos()

        if action == 'delete_produto':
            produto_id = request.POST.get('produto_id')
            produto = get_object_or_404(ProdutoSalao, id=produto_id)
            try:
                produto.delete()
                messages.success(request, 'Produto removido com sucesso.')
            except ProtectedError:
                messages.error(
                    request,
                    'Não foi possível remover: há movimentações de estoque vinculadas a esse produto.',
                )
            return _redirect_produtos()

    produtos_qs = _sort_produtos_por_codigo_natural(list(ProdutoSalao.objects.all()))
    edit_id = request.GET.get('edit')
    edit_produto = ProdutoSalao.objects.filter(id=edit_id).first() if edit_id else None

    context = {
        'active_tab': 'produtos',
        'produtos': produtos_qs,
        'edit_produto': edit_produto,
    }
    return render(request, 'salao/produtos.html', context)


@_salao_superuser_required
def estoque(request):
    ano, mes = _parse_competencia(request)
    produtos_ativos = _sort_produtos_por_codigo_natural(
        list(ProdutoSalao.objects.filter(ativo=True))
    )
    formas_pagamento_ativas = list(FormaPagamentoSalao.objects.filter(ativo=True).order_by('codigo'))

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'create_saida_estoque':
            raw_data = request.POST.get('data')
            try:
                ano_d, mes_d, dia_d = [int(part) for part in raw_data.split('-')]
                data_movimento = date(ano_d, mes_d, dia_d)
            except (TypeError, ValueError, AttributeError):
                messages.error(request, 'Data inválida para saída de estoque.')
                return _redirect_estoque(ano, mes)

            produto_id = request.POST.get('produto_id')
            produto = ProdutoSalao.objects.filter(id=produto_id, ativo=True).first()
            if not produto:
                messages.error(request, 'Selecione um produto ativo válido.')
                return _redirect_estoque(ano, mes)

            tipo_saida = (request.POST.get('tipo_saida') or '').strip().upper()
            if tipo_saida not in ('VENDA', 'USO_EM_CLIENTE'):
                messages.error(request, 'Selecione um tipo de saída válido.')
                return _redirect_estoque(ano, mes)

            quantidade = _parse_decimal(request.POST.get('quantidade') or '1', quantize_pattern='0.001')
            if quantidade is None or quantidade <= Decimal('0.000'):
                messages.error(request, 'Informe uma quantidade válida para saída.')
                return _redirect_estoque(ano, mes)

            observacao = (request.POST.get('observacao') or '').strip()

            with transaction.atomic():
                produto_locked = ProdutoSalao.objects.select_for_update().get(id=produto.id)
                if quantidade > produto_locked.saldo_atual:
                    messages.error(
                        request,
                        f'Estoque insuficiente para {produto_locked.codigo}. Saldo atual: {produto_locked.saldo_atual}.',
                    )
                    return _redirect_estoque(ano, mes)

                custo_unitario_aplicado = produto_locked.custo_medio_atual
                valor_custo_total = _quantize_money(quantidade * custo_unitario_aplicado)
                forma_pagamento = None
                parcelas = 1
                taxa_percentual = Decimal('0.00')
                valor_venda_unitario = None
                valor_bruto_venda = Decimal('0.00')
                valor_taxa = Decimal('0.00')
                valor_liquido_venda = Decimal('0.00')
                lucro_produto = Decimal('0.00')
                motivo = MovimentoEstoqueSalao.MOTIVO_USO_EM_CLIENTE

                if tipo_saida == 'VENDA':
                    forma_pagamento_id = request.POST.get('forma_pagamento_id')
                    forma_pagamento = FormaPagamentoSalao.objects.filter(
                        id=forma_pagamento_id, ativo=True
                    ).first()
                    if not forma_pagamento:
                        messages.error(request, 'Selecione uma forma de pagamento ativa para venda.')
                        return _redirect_estoque(ano, mes)

                    parcelas = _parse_parcelas(request.POST.get('parcelas'), default=1)
                    if not forma_pagamento.aceita_parcelamento:
                        parcelas = 1

                    taxa = TaxaFormaPagamentoSalao.objects.filter(
                        forma_pagamento=forma_pagamento,
                        parcelas=parcelas,
                    ).first()
                    taxa_percentual = taxa.percentual if taxa else Decimal('0.00')

                    valor_venda_unitario = _parse_decimal(request.POST.get('valor_venda_unitario'))
                    if valor_venda_unitario is None or valor_venda_unitario < Decimal('0.00'):
                        messages.error(request, 'Informe um valor de venda unitário válido.')
                        return _redirect_estoque(ano, mes)

                    valor_bruto_venda = _quantize_money(quantidade * valor_venda_unitario)
                    valor_taxa, valor_liquido_venda = _calcular_liquido_com_taxa(
                        valor_bruto_venda,
                        taxa_percentual,
                    )
                    lucro_produto = _quantize_money(valor_liquido_venda - valor_custo_total)
                    motivo = MovimentoEstoqueSalao.MOTIVO_VENDA

                produto_locked.saldo_atual = _quantize_quantity(produto_locked.saldo_atual - quantidade)
                produto_locked.save(update_fields=['saldo_atual', 'atualizado_em'])

                MovimentoEstoqueSalao.objects.create(
                    data=data_movimento,
                    produto=produto_locked,
                    tipo=MovimentoEstoqueSalao.TIPO_SAIDA,
                    motivo=motivo,
                    quantidade=quantidade,
                    custo_unitario_aplicado=custo_unitario_aplicado,
                    valor_custo_total=valor_custo_total,
                    valor_venda_unitario=valor_venda_unitario,
                    valor_bruto_venda=valor_bruto_venda,
                    taxa_percentual_aplicada=taxa_percentual,
                    valor_taxa=valor_taxa,
                    valor_liquido_venda=valor_liquido_venda,
                    lucro_produto=lucro_produto,
                    forma_pagamento=forma_pagamento,
                    parcelas=parcelas,
                    observacao=observacao,
                )

            if tipo_saida == 'VENDA':
                messages.success(
                    request,
                    f'Venda registrada. Líquido: R$ {valor_liquido_venda} | Lucro: R$ {lucro_produto}.',
                )
            else:
                messages.success(request, 'Saída de uso em cliente registrada com sucesso.')
            return _redirect_estoque(ano, mes)

    inicio_mes, fim_mes = _date_range_for_month(ano, mes)
    saidas_qs = (
        MovimentoEstoqueSalao.objects.filter(
            data__range=(inicio_mes, fim_mes),
            tipo=MovimentoEstoqueSalao.TIPO_SAIDA,
        )
        .select_related('produto', 'forma_pagamento')
        .order_by('-data', '-id')
    )
    saidas_mes = saidas_qs[:300]
    produtos_saldo = _sort_produtos_por_codigo_natural(
        list(ProdutoSalao.objects.filter(ativo=True))
    )
    produtos_alerta = [produto for produto in produtos_saldo if produto.saldo_atual <= produto.estoque_minimo]
    vendas_mes = saidas_qs.filter(motivo=MovimentoEstoqueSalao.MOTIVO_VENDA)
    resumo_vendas_mes = vendas_mes.aggregate(
        bruto=Sum('valor_bruto_venda'),
        taxas=Sum('valor_taxa'),
        liquido=Sum('valor_liquido_venda'),
        custo=Sum('valor_custo_total'),
        lucro=Sum('lucro_produto'),
    )

    context = {
        'active_tab': 'estoque',
        'ano': ano,
        'mes': mes,
        'month_options': MONTH_OPTIONS,
        'year_options': _build_year_options(),
        'data_padrao': date(ano, mes, min(date.today().day, calendar.monthrange(ano, mes)[1])),
        'produtos_ativos': produtos_ativos,
        'produtos_catalogo': [
            {
                'id': produto.id,
                'codigo': produto.codigo,
                'nome': produto.nome,
                'valor_venda_padrao': str(produto.valor_venda_padrao),
            }
            for produto in produtos_ativos
        ],
        'formas_pagamento_ativas': formas_pagamento_ativas,
        'formas_catalogo': _build_formas_catalogo(formas_pagamento_ativas),
        'saidas_mes': saidas_mes,
        'produtos_saldo': produtos_saldo,
        'produtos_alerta': produtos_alerta,
        'resumo_vendas_mes': {
            'bruto': resumo_vendas_mes['bruto'] or Decimal('0.00'),
            'taxas': resumo_vendas_mes['taxas'] or Decimal('0.00'),
            'liquido': resumo_vendas_mes['liquido'] or Decimal('0.00'),
            'custo': resumo_vendas_mes['custo'] or Decimal('0.00'),
            'lucro': resumo_vendas_mes['lucro'] or Decimal('0.00'),
        },
    }
    return render(request, 'salao/estoque.html', context)


@_salao_superuser_required
def pagamentos(request):
    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'create_forma_pagamento':
            codigo = _normalize_codigo(request.POST.get('codigo'))
            nome = (request.POST.get('nome') or '').strip()
            aceita_parcelamento = _parse_checkbox(request.POST.get('aceita_parcelamento'))
            ativo = _parse_checkbox(request.POST.get('ativo'))

            if not codigo or not nome:
                messages.error(request, 'Informe código e nome da forma de pagamento.')
                return _redirect_pagamentos()

            if FormaPagamentoSalao.objects.filter(codigo=codigo).exists():
                messages.error(request, f'Já existe forma de pagamento com código {codigo}.')
                return _redirect_pagamentos()

            FormaPagamentoSalao.objects.create(
                codigo=codigo,
                nome=nome,
                aceita_parcelamento=aceita_parcelamento,
                ativo=ativo,
            )
            messages.success(request, 'Forma de pagamento criada com sucesso.')
            return _redirect_pagamentos()

        if action == 'update_forma_pagamento':
            forma_id = request.POST.get('forma_id')
            forma = get_object_or_404(FormaPagamentoSalao, id=forma_id)

            codigo = _normalize_codigo(request.POST.get('codigo'))
            nome = (request.POST.get('nome') or '').strip()
            aceita_parcelamento = _parse_checkbox(request.POST.get('aceita_parcelamento'))
            ativo = _parse_checkbox(request.POST.get('ativo'))

            if not codigo or not nome:
                messages.error(request, 'Informe código e nome da forma de pagamento.')
                return _redirect_pagamentos()

            if FormaPagamentoSalao.objects.exclude(id=forma.id).filter(codigo=codigo).exists():
                messages.error(request, f'Já existe outra forma de pagamento com código {codigo}.')
                return _redirect_pagamentos()

            forma.codigo = codigo
            forma.nome = nome
            forma.aceita_parcelamento = aceita_parcelamento
            forma.ativo = ativo
            forma.save()
            messages.success(request, 'Forma de pagamento atualizada com sucesso.')
            return _redirect_pagamentos(forma_taxa_id=forma.id)

        if action == 'delete_forma_pagamento':
            forma_id = request.POST.get('forma_id')
            forma = get_object_or_404(FormaPagamentoSalao, id=forma_id)
            try:
                forma.delete()
                messages.success(request, 'Forma de pagamento removida com sucesso.')
            except ProtectedError:
                messages.error(
                    request,
                    'Não foi possível remover: existem lançamentos vinculados a essa forma de pagamento.',
                )
            return _redirect_pagamentos()

        if action == 'save_taxas_forma':
            forma_id = request.POST.get('forma_id')
            forma = get_object_or_404(FormaPagamentoSalao, id=forma_id)

            parcelas_range = range(1, 13) if forma.aceita_parcelamento else range(1, 2)
            taxas_to_upsert = []
            parcelas_to_keep = set()

            for parcela in parcelas_range:
                raw_percentual = (request.POST.get(f'taxa_{parcela}') or '').strip()
                if raw_percentual == '':
                    continue
                percentual = _parse_decimal(raw_percentual)
                if percentual is None or percentual < Decimal('0.00') or percentual > Decimal('100.00'):
                    messages.error(request, f'Taxa inválida para {parcela}x.')
                    return _redirect_pagamentos(forma_taxa_id=forma.id)
                parcelas_to_keep.add(parcela)
                taxas_to_upsert.append((parcela, percentual))

            with transaction.atomic():
                for parcela, percentual in taxas_to_upsert:
                    TaxaFormaPagamentoSalao.objects.update_or_create(
                        forma_pagamento=forma,
                        parcelas=parcela,
                        defaults={'percentual': percentual},
                    )
                TaxaFormaPagamentoSalao.objects.filter(forma_pagamento=forma).exclude(
                    parcelas__in=parcelas_to_keep
                ).delete()

            messages.success(request, 'Taxas salvas com sucesso.')
            return _redirect_pagamentos(forma_taxa_id=forma.id)

    formas_qs = list(FormaPagamentoSalao.objects.all().order_by('codigo'))
    edit_forma = None
    edit_forma_id = request.GET.get('edit_forma')
    if edit_forma_id:
        edit_forma = FormaPagamentoSalao.objects.filter(id=edit_forma_id).first()

    forma_taxa = None
    forma_taxa_id = request.GET.get('forma_taxa')
    if forma_taxa_id:
        forma_taxa = FormaPagamentoSalao.objects.filter(id=forma_taxa_id).first()
    if not forma_taxa and formas_qs:
        forma_taxa = formas_qs[0]

    taxas_map = {}
    parcelas_range = range(1, 2)
    if forma_taxa:
        parcelas_range = range(1, 13) if forma_taxa.aceita_parcelamento else range(1, 2)
        taxas_map = {
            taxa.parcelas: taxa.percentual
            for taxa in TaxaFormaPagamentoSalao.objects.filter(forma_pagamento=forma_taxa)
        }
    taxas_rows = [
        {
            'parcela': parcela,
            'percentual': taxas_map.get(parcela),
        }
        for parcela in parcelas_range
    ]

    context = {
        'active_tab': 'pagamentos',
        'formas_pagamento': formas_qs,
        'edit_forma': edit_forma,
        'forma_taxa': forma_taxa,
        'taxas_rows': taxas_rows,
    }
    return render(request, 'salao/pagamentos.html', context)


@_salao_superuser_required
def dashboard(request):
    ano, mes = _parse_competencia(request)

    comissao, _ = ComissaoMensalSalao.objects.get_or_create(
        ano=ano,
        mes=mes,
        defaults={'percentual': Decimal('20.00')},
    )

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'update_meta':
            raw_meta = (request.POST.get('meta_faturamento') or '').strip()
            if raw_meta == '':
                comissao.meta_faturamento = None
                comissao.save(update_fields=['meta_faturamento', 'updated_at'])
                messages.success(request, 'Meta mensal removida.')
                return _redirect_dashboard(ano, mes)

            meta_value = _parse_decimal(raw_meta)
            if meta_value is None or meta_value < Decimal('0.00'):
                messages.error(request, 'Informe uma meta mensal válida.')
                return _redirect_dashboard(ano, mes)

            comissao.meta_faturamento = meta_value
            comissao.save(update_fields=['meta_faturamento', 'updated_at'])
            messages.success(request, 'Meta mensal atualizada.')
            return _redirect_dashboard(ano, mes)

    lancamentos_mes = LancamentoSalao.objects.filter(data__year=ano, data__month=mes)
    lancamentos_normais_mes = lancamentos_mes.filter(permuta=False, sem_comissao=False)
    lancamentos_permuta_mes = lancamentos_mes.filter(permuta=True)
    lancamentos_sem_comissao_mes = lancamentos_mes.filter(permuta=False, sem_comissao=True)
    despesas_mes = DespesaSalao.objects.filter(data__year=ano, data__month=mes)
    vendas_produto_mes = MovimentoEstoqueSalao.objects.filter(
        data__year=ano,
        data__month=mes,
        tipo=MovimentoEstoqueSalao.TIPO_SAIDA,
        motivo=MovimentoEstoqueSalao.MOTIVO_VENDA,
    )

    faturamento_liquido = (
        lancamentos_normais_mes.aggregate(total=Sum('valor_cobrado'))['total'] or Decimal('0.00')
    )
    faturamento_bruto_cliente = (
        lancamentos_normais_mes.aggregate(total=Sum('valor_bruto'))['total'] or Decimal('0.00')
    )
    taxas_total = lancamentos_normais_mes.aggregate(total=Sum('valor_taxa'))['total'] or Decimal('0.00')
    permuta_total_mes = (
        lancamentos_permuta_mes.aggregate(total=Sum('valor_bruto'))['total'] or Decimal('0.00')
    )
    atendimentos_permuta_total = lancamentos_permuta_mes.count()
    sem_comissao_total_mes = (
        lancamentos_sem_comissao_mes.aggregate(total=Sum('valor_cobrado'))['total'] or Decimal('0.00')
    )
    atendimentos_sem_comissao_total = lancamentos_sem_comissao_mes.count()
    atendimentos_normais_total = lancamentos_normais_mes.count()
    despesas_total = despesas_mes.aggregate(total=Sum('valor'))['total'] or Decimal('0.00')
    vendas_produto_brutas = (
        vendas_produto_mes.aggregate(total=Sum('valor_bruto_venda'))['total'] or Decimal('0.00')
    )
    taxas_produto_total = vendas_produto_mes.aggregate(total=Sum('valor_taxa'))['total'] or Decimal('0.00')
    vendas_produto_liquidas = (
        vendas_produto_mes.aggregate(total=Sum('valor_liquido_venda'))['total'] or Decimal('0.00')
    )
    custo_produto_vendido = (
        vendas_produto_mes.aggregate(total=Sum('valor_custo_total'))['total'] or Decimal('0.00')
    )
    lucro_produto = vendas_produto_mes.aggregate(total=Sum('lucro_produto'))['total'] or Decimal('0.00')
    atendimentos_total = lancamentos_mes.count()
    ticket_medio = (
        (faturamento_liquido / Decimal(atendimentos_normais_total)).quantize(Decimal('0.01'))
        if atendimentos_normais_total > 0
        else Decimal('0.00')
    )
    ticket_permuta = (
        (permuta_total_mes / Decimal(atendimentos_permuta_total)).quantize(Decimal('0.01'))
        if atendimentos_permuta_total > 0
        else Decimal('0.00')
    )
    ticket_sem_comissao = (
        (sem_comissao_total_mes / Decimal(atendimentos_sem_comissao_total)).quantize(Decimal('0.01'))
        if atendimentos_sem_comissao_total > 0
        else Decimal('0.00')
    )

    comissao_calculada = (
        (faturamento_liquido * comissao.percentual) / Decimal('100.00')
    ).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

    comissao_paga = comissao_calculada
    valor_pos_comissao = faturamento_liquido - comissao_paga
    lucro = valor_pos_comissao - despesas_total
    impacto_taxas_percentual = (
        (taxas_total / faturamento_bruto_cliente) * Decimal('100.00')
    ).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP) if faturamento_bruto_cliente > Decimal('0.00') else Decimal('0.00')

    meta_faturamento = comissao.meta_faturamento
    percentual_meta_atingido = None
    valor_faltante_meta = None
    if meta_faturamento and meta_faturamento > Decimal('0.00'):
        percentual_meta_atingido = (
            (faturamento_liquido / meta_faturamento) * Decimal('100.00')
        ).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        valor_faltante_meta = max(
            (meta_faturamento - faturamento_liquido).quantize(
                Decimal('0.01'),
                rounding=ROUND_HALF_UP,
            ),
            Decimal('0.00'),
        )

    ranking_servicos = (
        lancamentos_normais_mes.values('servico__codigo', 'servico__nome')
        .annotate(quantidade=Count('id'), total=Sum('valor_cobrado'))
        .order_by('-quantidade', '-total', 'servico__codigo')
    )

    despesas_por_categoria = (
        despesas_mes.values('categoria__nome')
        .annotate(total=Sum('valor'))
        .order_by('-total', 'categoria__nome')
    )
    despesas_por_subcategoria = (
        despesas_mes.filter(subcategoria__isnull=False).annotate(
            subcategoria_nome=Coalesce(
                'subcategoria__nome',
                Value('', output_field=CharField()),
            )
        )
        .values('subcategoria_nome')
        .annotate(total=Sum('valor'))
        .order_by('-total', 'subcategoria_nome')
    )
    produtos_alerta = (
        ProdutoSalao.objects.filter(ativo=True, saldo_atual__lte=F('estoque_minimo'))
        .order_by('saldo_atual', 'codigo')[:15]
    )

    serie_meses = _iter_months_backwards(ano, mes, quantidade=6)
    existing_commissions = {
        (item.ano, item.mes): item.percentual
        for item in ComissaoMensalSalao.objects.filter(
            ano__in={item_ano for item_ano, _ in serie_meses}
        )
    }

    chart_labels = []
    chart_faturamento = []
    chart_taxas = []
    chart_despesas = []
    chart_lucro = []
    chart_atendimentos = []
    chart_ticket = []
    chart_permuta = []

    for ano_item, mes_item in serie_meses:
        label = f"{mes_item:02d}/{ano_item}"
        lancamentos_item = LancamentoSalao.objects.filter(data__year=ano_item, data__month=mes_item)
        lancamentos_item_normais = lancamentos_item.filter(permuta=False, sem_comissao=False)
        lancamentos_item_permuta = lancamentos_item.filter(permuta=True)
        despesas_item = DespesaSalao.objects.filter(data__year=ano_item, data__month=mes_item)

        faturamento_item = (
            lancamentos_item_normais.aggregate(total=Sum('valor_cobrado'))['total'] or Decimal('0.00')
        )
        taxas_item = lancamentos_item_normais.aggregate(total=Sum('valor_taxa'))['total'] or Decimal('0.00')
        permuta_item = (
            lancamentos_item_permuta.aggregate(total=Sum('valor_bruto'))['total'] or Decimal('0.00')
        )
        despesas_item_total = despesas_item.aggregate(total=Sum('valor'))['total'] or Decimal('0.00')
        percentual_item = existing_commissions.get((ano_item, mes_item), Decimal('20.00'))
        comissao_item = (faturamento_item * percentual_item / Decimal('100.00')).quantize(
            Decimal('0.01'),
            rounding=ROUND_HALF_UP,
        )
        lucro_item = faturamento_item - comissao_item - despesas_item_total
        qtd_item = lancamentos_item.count()
        qtd_item_normais = lancamentos_item_normais.count()
        ticket_item = (
            (faturamento_item / Decimal(qtd_item_normais)).quantize(Decimal('0.01'))
            if qtd_item_normais > 0
            else Decimal('0.00')
        )

        chart_labels.append(label)
        chart_faturamento.append(float(faturamento_item))
        chart_taxas.append(float(taxas_item))
        chart_despesas.append(float(despesas_item_total))
        chart_lucro.append(float(lucro_item))
        chart_atendimentos.append(qtd_item)
        chart_ticket.append(float(ticket_item))
        chart_permuta.append(float(permuta_item))

    atendimentos_por_dia_qs = (
        lancamentos_mes.values('data__day')
        .annotate(qtd=Count('id'), total=Sum('valor_cobrado'))
        .order_by('data__day')
    )
    map_qtd_por_dia = {item['data__day']: item['qtd'] for item in atendimentos_por_dia_qs}
    map_total_por_dia = {item['data__day']: float(item['total'] or 0) for item in atendimentos_por_dia_qs}
    dias_mes_chart = calendar.monthrange(ano, mes)[1]
    dias_labels = [f"{dia:02d}" for dia in range(1, dias_mes_chart + 1)]
    dias_qtd = [map_qtd_por_dia.get(dia, 0) for dia in range(1, dias_mes_chart + 1)]
    dias_total = [map_total_por_dia.get(dia, 0.0) for dia in range(1, dias_mes_chart + 1)]

    context = {
        'active_tab': 'dashboard',
        'ano': ano,
        'mes': mes,
        'month_options': MONTH_OPTIONS,
        'year_options': _build_year_options(),
        'faturamento_bruto': faturamento_liquido,
        'faturamento_bruto_cliente': faturamento_bruto_cliente,
        'taxas_total': taxas_total,
        'impacto_taxas_percentual': impacto_taxas_percentual,
        'comissao_percentual': comissao.percentual,
        'comissao_calculada': comissao_calculada,
        'comissao_paga': comissao_paga,
        'valor_pos_comissao': valor_pos_comissao,
        'despesas_total': despesas_total,
        'lucro': lucro,
        'meta_faturamento': meta_faturamento,
        'percentual_meta_atingido': percentual_meta_atingido,
        'valor_faltante_meta': valor_faltante_meta,
        'atendimentos_total': atendimentos_total,
        'ticket_medio': ticket_medio,
        'ticket_permuta': ticket_permuta,
        'atendimentos_permuta_total': atendimentos_permuta_total,
        'permuta_total_mes': permuta_total_mes,
        'ticket_sem_comissao': ticket_sem_comissao,
        'atendimentos_sem_comissao_total': atendimentos_sem_comissao_total,
        'sem_comissao_total_mes': sem_comissao_total_mes,
        'vendas_produto_brutas': vendas_produto_brutas,
        'taxas_produto_total': taxas_produto_total,
        'vendas_produto_liquidas': vendas_produto_liquidas,
        'custo_produto_vendido': custo_produto_vendido,
        'lucro_produto': lucro_produto,
        'produtos_alerta': produtos_alerta,
        'ranking_servicos': ranking_servicos,
        'despesas_por_categoria': despesas_por_categoria,
        'despesas_por_subcategoria': despesas_por_subcategoria,
        'meta_bullet_chart': {
            'tem_meta': bool(meta_faturamento and meta_faturamento > Decimal('0.00')),
            'meta': float(meta_faturamento or Decimal('0.00')),
            'realizado': float(faturamento_liquido),
            'percentual_real': float(percentual_meta_atingido or Decimal('0.00')),
            'faltante': float(valor_faltante_meta or Decimal('0.00')),
        },
        'comparativo_chart': {
            'labels': chart_labels,
            'faturamento': chart_faturamento,
            'taxas': chart_taxas,
            'despesas': chart_despesas,
            'lucro': chart_lucro,
        },
        'operacao_chart': {
            'labels': chart_labels,
            'atendimentos': chart_atendimentos,
            'ticket_medio': chart_ticket,
        },
        'permuta_chart': {
            'labels': chart_labels,
            'permuta': chart_permuta,
        },
        'atendimentos_dia_chart': {
            'labels': dias_labels,
            'atendimentos': dias_qtd,
            'faturamento': dias_total,
        },
    }
    return render(request, 'salao/dashboard.html', context)


@_salao_superuser_required
def dashboard_relatorio_lancamentos(request):
    ano, mes = _parse_competencia(request)
    comissao_percentual = Decimal('20.00')
    fator_pos_comissao = (Decimal('100.00') - comissao_percentual) / Decimal('100.00')

    lancamentos_mes = (
        LancamentoSalao.objects.filter(data__year=ano, data__month=mes)
        .select_related('servico', 'forma_pagamento')
        .order_by('data', 'id')
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"Lancamentos {mes:02d}-{ano}"

    headers = [
        'Data',
        'Servico',
        'Tipo',
        'Forma de pagamento',
        'Valor',
        'Parcelas',
        'Taxa (%)',
        'Valor taxa',
        'Valor liquido',
        'Valor 20%',
        'Valor apos 20%',
    ]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    total_valor = Decimal('0.00')
    total_taxa = Decimal('0.00')
    total_liquido = Decimal('0.00')
    total_20 = Decimal('0.00')
    total_pos_20 = Decimal('0.00')

    for lancamento in lancamentos_mes:
        valor = lancamento.valor_bruto or Decimal('0.00')
        valor_taxa = lancamento.valor_taxa or Decimal('0.00')
        valor_liquido = lancamento.valor_cobrado or Decimal('0.00')
        if lancamento.permuta or lancamento.sem_comissao:
            valor_20 = Decimal('0.00')
            valor_pos_20 = valor_liquido
        else:
            valor_20 = (valor_liquido * comissao_percentual / Decimal('100.00')).quantize(
                Decimal('0.01'),
                rounding=ROUND_HALF_UP,
            )
            valor_pos_20 = (valor_liquido * fator_pos_comissao).quantize(
                Decimal('0.01'),
                rounding=ROUND_HALF_UP,
            )
        total_valor += valor
        total_taxa += valor_taxa
        total_liquido += valor_liquido
        total_20 += valor_20
        total_pos_20 += valor_pos_20
        sheet.append(
            [
                lancamento.data.strftime('%d/%m/%Y'),
                lancamento.servico.nome if lancamento.servico else '',
                'PERMUTA' if lancamento.permuta else ('SEM COMISSAO' if lancamento.sem_comissao else 'NORMAL'),
                lancamento.forma_pagamento.nome if lancamento.forma_pagamento else 'Nao informado',
                float(valor),
                int(lancamento.parcelas or 1),
                float(lancamento.taxa_percentual_aplicada or Decimal('0.00')),
                float(valor_taxa),
                float(valor_liquido),
                float(valor_20),
                float(valor_pos_20),
            ]
        )

    total_row_idx = sheet.max_row + 2
    sheet.cell(row=total_row_idx, column=1, value='TOTAL')
    sheet.cell(row=total_row_idx, column=5, value=float(total_valor))
    sheet.cell(row=total_row_idx, column=8, value=float(total_taxa))
    sheet.cell(row=total_row_idx, column=9, value=float(total_liquido))
    sheet.cell(row=total_row_idx, column=10, value=float(total_20))
    sheet.cell(row=total_row_idx, column=11, value=float(total_pos_20))
    for col in (1, 5, 8, 9, 10, 11):
        sheet.cell(row=total_row_idx, column=col).font = Font(bold=True)

    sheet.column_dimensions['A'].width = 13
    sheet.column_dimensions['B'].width = 26
    sheet.column_dimensions['C'].width = 12
    sheet.column_dimensions['D'].width = 22
    sheet.column_dimensions['E'].width = 13
    sheet.column_dimensions['F'].width = 10
    sheet.column_dimensions['G'].width = 10
    sheet.column_dimensions['H'].width = 13
    sheet.column_dimensions['I'].width = 13
    sheet.column_dimensions['J'].width = 13
    sheet.column_dimensions['K'].width = 16

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="relatorio_lancamentos_{ano}_{mes:02d}.xlsx"'
    )
    workbook.save(response)
    return response


@_salao_superuser_required
def conferencia(request):
    ano, mes = _parse_competencia(request)

    if request.method == 'POST':
        action = request.POST.get('action')
        retorno = (request.POST.get('retorno') or '').strip()

        if action == 'toggle_conferido':
            conferido = _parse_checkbox(request.POST.get('conferido'))
            origem = _marcar_conferido(request.POST.get('item'), conferido)
            if origem is None:
                return JsonResponse({'ok': False}, status=400)
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'ok': True, 'conferido': origem.conferido})
            return _redirect_conferencia(ano, mes, retorno=retorno)

        if action == 'ajustar_lancamento':
            lancamento = get_object_or_404(LancamentoSalao, pk=request.POST.get('lancamento_id'))
            dia_origem = lancamento.data.day
            valor_bruto = _parse_decimal(request.POST.get('valor_bruto'))
            if valor_bruto is None or valor_bruto < Decimal('0.00'):
                messages.error(request, 'Informe um valor bruto válido.')
                return _redirect_conferencia(ano, mes, retorno=retorno)

            forma_pagamento = FormaPagamentoSalao.objects.filter(
                pk=request.POST.get('forma_pagamento_id'),
                ativo=True,
            ).first()
            if not forma_pagamento:
                messages.error(request, 'Selecione uma forma de pagamento ativa.')
                return _redirect_conferencia(ano, mes, retorno=retorno)

            parcelas = _parse_parcelas(request.POST.get('parcelas'), default=1)
            if not forma_pagamento.aceita_parcelamento:
                parcelas = 1

            taxa = TaxaFormaPagamentoSalao.objects.filter(
                forma_pagamento=forma_pagamento,
                parcelas=parcelas,
            ).first()
            if not taxa:
                messages.error(
                    request,
                    f'Taxa não cadastrada para {forma_pagamento.nome} em {parcelas}x.',
                )
                return _redirect_conferencia(ano, mes, retorno=retorno)

            valor_taxa, valor_liquido = _calcular_liquido_com_taxa(valor_bruto, taxa.percentual)
            lancamento.valor_bruto = valor_bruto
            lancamento.forma_pagamento = forma_pagamento
            lancamento.parcelas = parcelas
            lancamento.taxa_percentual_aplicada = taxa.percentual
            lancamento.valor_taxa = valor_taxa
            lancamento.valor_cobrado = valor_liquido
            lancamento.save(update_fields=[
                'valor_bruto',
                'forma_pagamento',
                'parcelas',
                'taxa_percentual_aplicada',
                'valor_taxa',
                'valor_cobrado',
                'atualizado_em',
            ])
            messages.success(
                request,
                f'Lançamento ajustado para R$ {valor_bruto} em {forma_pagamento.nome} '
                f'{parcelas}x. Líquido: R$ {valor_liquido}.',
            )
            return _redirect_conferencia(ano, mes, retorno=retorno)

        if action in {'conferir_dia', 'desconferir_dia'}:
            dia = _parse_day(request, ano, mes)
            if dia is None:
                messages.error(request, 'Dia inválido.')
                return _redirect_conferencia(ano, mes, None)
            marcar = action == 'conferir_dia'
            valores = {
                'conferido': marcar,
                'conferido_em': timezone.now() if marcar else None,
            }
            data_alvo = date(ano, mes, dia)
            atualizados = LancamentoSalao.objects.filter(
                data=data_alvo, conferido=not marcar,
            ).update(**valores)
            atualizados += MovimentoEstoqueSalao.objects.filter(
                data=data_alvo,
                tipo=MovimentoEstoqueSalao.TIPO_SAIDA,
                motivo=MovimentoEstoqueSalao.MOTIVO_VENDA,
                conferido=not marcar,
            ).update(**valores)
            verbo = 'conferidos' if marcar else 'desmarcados'
            messages.success(request, f'{atualizados} item(ns) {verbo} no dia {dia:02d}/{mes:02d}.')
            return _redirect_conferencia(ano, mes, retorno=retorno)

        if action == 'importar_extrato':
            arquivo = request.FILES.get('extrato')
            if not arquivo:
                messages.error(request, 'Selecione o arquivo CSV do relatório da adquirente.')
                return _redirect_conferencia(ano, mes, None)
            try:
                transacoes = _parse_extrato_csv(arquivo)
            except ValueError as erro:
                messages.error(request, str(erro))
                return _redirect_conferencia(ano, mes, None)

            if not transacoes:
                messages.error(request, 'Nenhuma transação encontrada no arquivo.')
                return _redirect_conferencia(ano, mes, None)

            request.session[CONFERENCIA_SESSION_KEY] = transacoes
            competencias = Counter((t['ano'], t['mes']) for t in transacoes)
            resumo = ', '.join(
                f'{item_mes:02d}/{item_ano} ({quantidade})'
                for (item_ano, item_mes), quantidade in sorted(competencias.items())
            )
            messages.success(
                request,
                f'Extrato importado: {len(transacoes)} transação(ões) — {resumo}.',
            )

            # O arquivo é que manda: se a competência aberta não tem nada dele,
            # abre direto o mês com mais transações em vez de mostrar tela vazia.
            if (ano, mes) not in competencias:
                (ano, mes), _ = competencias.most_common(1)[0]
            return _redirect_conferencia(ano, mes, None)

        if action == 'ignorar_transacao':
            identificador = (request.POST.get('identificador') or '').strip()
            if identificador:
                TransacaoIgnoradaSalao.objects.get_or_create(
                    identificador=identificador,
                    defaults={'referencia': (request.POST.get('referencia') or '')[:200]},
                )
                messages.success(request, 'Transação marcada como fora do salão.')
            return _redirect_conferencia(ano, mes, retorno=retorno)

        if action == 'reativar_transacao':
            TransacaoIgnoradaSalao.objects.filter(
                identificador=(request.POST.get('identificador') or '').strip()
            ).delete()
            messages.success(request, 'Transação voltou para a conferência.')
            return _redirect_conferencia(ano, mes, retorno=retorno)

        if action == 'limpar_extrato':
            request.session.pop(CONFERENCIA_SESSION_KEY, None)
            messages.success(request, 'Extrato removido da conferência.')
            return _redirect_conferencia(ano, mes, None)

    dia_filtro = None
    raw_dia = (request.GET.get('dia') or '').strip()
    if raw_dia:
        dia_filtro = _parse_day(request, ano, mes)

    somente_pendentes = request.GET.get('pendentes') == '1'
    tolerancia = _parse_decimal(request.GET.get('tolerancia'))
    if tolerancia is None or tolerancia < Decimal('0.00'):
        tolerancia = Decimal('1.00')

    transacoes_arquivo = request.session.get(CONFERENCIA_SESSION_KEY) or []
    transacoes_sessao = [
        t for t in transacoes_arquivo if t['ano'] == ano and t['mes'] == mes
    ]
    identificadores_ignorados = set(
        TransacaoIgnoradaSalao.objects.values_list('identificador', flat=True)
    )

    transacoes_aprovadas = []
    transacoes_recusadas = []
    transacoes_ignoradas = []
    for transacao in transacoes_sessao:
        if transacao['identificador'] in identificadores_ignorados:
            transacoes_ignoradas.append(transacao)
        elif _normalizar_texto(transacao['status']) == 'aprovada':
            transacoes_aprovadas.append(transacao)
        else:
            transacoes_recusadas.append(transacao)

    # Serviços e vendas de produto caem no mesmo extrato, então conferem juntos.
    itens_mes = _itens_conferencia_do_mes(ano, mes)

    transacoes_por_dia = defaultdict(list)
    for transacao in transacoes_aprovadas:
        transacoes_por_dia[transacao['dia']].append(transacao)

    itens_por_dia = defaultdict(list)
    for item in itens_mes:
        itens_por_dia[item.data.day].append(item)

    # Passada 1: cada dia contra as transações do próprio dia.
    resultado_por_dia = {}
    pendentes_por_dia = defaultdict(list)
    disponiveis_por_dia = defaultdict(list)
    for dia in sorted(set(itens_por_dia) | set(transacoes_por_dia)):
        do_dia = itens_por_dia.get(dia, [])
        # Permuta não passa pelo banco: fica de fora do pareamento, mas continua visível.
        conciliaveis = [i for i in do_dia if not i.permuta]
        pares, sem_par_sistema, sem_par_extrato = _parear_dia(
            conciliaveis,
            transacoes_por_dia.get(dia, []),
            tolerancia=tolerancia,
        )
        resultado_por_dia[dia] = {
            'pares': pares,
            'permutas': [i for i in do_dia if i.permuta],
            'conciliaveis': conciliaveis,
        }
        pendentes_por_dia[dia] = sem_par_sistema
        disponiveis_por_dia[dia] = sem_par_extrato

    # Passada 2: sobras que só fecham cruzando dias (pagamento junto de dois atendimentos).
    for par in _parear_entre_dias(pendentes_por_dia, disponiveis_por_dia, tolerancia):
        resultado_por_dia[par['dia_transacao']]['pares'].append(par)

    dias = sorted(resultado_por_dia)
    if dia_filtro is not None:
        dias = [d for d in dias if d == dia_filtro]

    dias_context = []
    total_sistema = Decimal('0.00')
    total_extrato = Decimal('0.00')
    total_conferidos = 0
    total_itens = 0
    total_sem_par_sistema = 0
    total_sem_par_extrato = 0

    for dia in dias:
        resultado = resultado_por_dia[dia]
        do_dia = itens_por_dia.get(dia, [])
        sem_par_sistema = pendentes_por_dia[dia]
        sem_par_extrato = disponiveis_por_dia[dia]

        soma_sistema = sum((i.valor_bruto for i in resultado['conciliaveis']), Decimal('0.00'))
        soma_extrato = sum(
            (Decimal(t['bruto']) for t in transacoes_por_dia.get(dia, [])),
            Decimal('0.00'),
        )
        conferidos_dia = sum(1 for i in do_dia if i.conferido)

        total_sistema += soma_sistema
        total_extrato += soma_extrato
        total_conferidos += conferidos_dia
        total_itens += len(do_dia)
        total_sem_par_sistema += len(sem_par_sistema)
        total_sem_par_extrato += len(sem_par_extrato)

        if somente_pendentes and conferidos_dia == len(do_dia) and not sem_par_extrato:
            continue

        dias_context.append({
            'dia': dia,
            'data': date(ano, mes, dia),
            'pares': sorted(resultado['pares'], key=lambda p: p['lancamentos'][0].data),
            'sem_par_sistema': sem_par_sistema,
            'sem_par_extrato': sem_par_extrato,
            'permutas': resultado['permutas'],
            'quantidade': len(do_dia),
            'conferidos': conferidos_dia,
            'pendentes': len(do_dia) - conferidos_dia,
            'total_sistema': soma_sistema,
            'total_extrato': soma_extrato,
            'diferenca': (soma_sistema - soma_extrato).quantize(Decimal('0.01')),
            'tudo_conferido': len(do_dia) > 0 and conferidos_dia == len(do_dia),
        })

    context = {
        'active_tab': 'conferencia',
        'ano': ano,
        'mes': mes,
        'month_options': MONTH_OPTIONS,
        'year_options': _build_year_options(),
        'dia_filtro': dia_filtro,
        'somente_pendentes': somente_pendentes,
        'dias': dias_context,
        'tolerancia': tolerancia,
        'formas_pagamento_ativas': FormaPagamentoSalao.objects.filter(ativo=True).order_by('codigo'),
        'tem_extrato': bool(transacoes_sessao),
        'extrato_outra_competencia': [
            {'ano': item_ano, 'mes': item_mes, 'quantidade': quantidade}
            for (item_ano, item_mes), quantidade in sorted(
                Counter((t['ano'], t['mes']) for t in transacoes_arquivo).items()
            )
        ] if transacoes_arquivo and not transacoes_sessao else [],
        'transacoes_recusadas': transacoes_recusadas,
        'transacoes_ignoradas': transacoes_ignoradas,
        'resumo': {
            'total_sistema': total_sistema,
            'total_extrato': total_extrato,
            'diferenca': (total_sistema - total_extrato).quantize(Decimal('0.01')),
            'lancamentos': total_itens,
            'conferidos': total_conferidos,
            'pendentes': total_itens - total_conferidos,
            'sem_par_sistema': total_sem_par_sistema,
            'sem_par_extrato': total_sem_par_extrato,
        },
    }
    return render(request, 'salao/conferencia.html', context)


@_salao_superuser_required
def grid_lancamentos(request):
    ano, mes = _parse_competencia(request)
    servico_id = (request.GET.get('servico_id') or '').strip()
    forma_pagamento_id = (request.GET.get('forma_pagamento_id') or '').strip()

    lancamentos_qs = (
        LancamentoSalao.objects.filter(data__year=ano, data__month=mes)
        .select_related('servico', 'forma_pagamento')
    )
    if servico_id:
        lancamentos_qs = lancamentos_qs.filter(servico_id=servico_id)
    if forma_pagamento_id:
        lancamentos_qs = lancamentos_qs.filter(forma_pagamento_id=forma_pagamento_id)

    resumo_totais = lancamentos_qs.aggregate(
        bruto=Sum('valor_bruto'),
        taxa=Sum('valor_taxa'),
        liquido=Sum('valor_cobrado'),
        quantidade=Count('id'),
    )
    # Mesma base do dashboard (exclui permuta e sem comissão), para os totais baterem entre as telas.
    resumo_normais = lancamentos_qs.filter(permuta=False, sem_comissao=False).aggregate(
        bruto=Sum('valor_bruto'),
        liquido=Sum('valor_cobrado'),
        quantidade=Count('id'),
    )
    lancamentos_qs = lancamentos_qs.order_by('-data', '-id')
    paginator = Paginator(lancamentos_qs, 50)
    page_obj = paginator.get_page(request.GET.get('page') or 1)
    query_params = request.GET.copy()
    query_params.pop('page', None)

    context = {
        'active_tab': 'dashboard',
        'ano': ano,
        'mes': mes,
        'month_options': MONTH_OPTIONS,
        'year_options': _build_year_options(),
        'servicos_ativos': ServicoSalao.objects.filter(ativo=True).order_by('codigo'),
        'formas_pagamento_ativas': FormaPagamentoSalao.objects.filter(ativo=True).order_by('codigo'),
        'filtro_servico_id': servico_id,
        'filtro_forma_pagamento_id': forma_pagamento_id,
        'page_obj': page_obj,
        'pagination_query': query_params.urlencode(),
        'totais': {
            'bruto': resumo_totais['bruto'] or Decimal('0.00'),
            'taxa': resumo_totais['taxa'] or Decimal('0.00'),
            'liquido': resumo_totais['liquido'] or Decimal('0.00'),
            'quantidade': resumo_totais['quantidade'] or 0,
            'bruto_normais': resumo_normais['bruto'] or Decimal('0.00'),
            'liquido_normais': resumo_normais['liquido'] or Decimal('0.00'),
            'quantidade_normais': resumo_normais['quantidade'] or 0,
        },
    }
    return render(request, 'salao/grid_lancamentos.html', context)


@_salao_superuser_required
def grid_despesas(request):
    ano, mes = _parse_competencia(request)
    categoria_id = (request.GET.get('categoria_id') or '').strip()
    subcategoria_id = (request.GET.get('subcategoria_id') or '').strip()

    despesas_qs = DespesaSalao.objects.filter(data__year=ano, data__month=mes).select_related(
        'categoria', 'subcategoria'
    )
    if categoria_id:
        despesas_qs = despesas_qs.filter(categoria_id=categoria_id)
    if subcategoria_id:
        despesas_qs = despesas_qs.filter(subcategoria_id=subcategoria_id)

    resumo_totais = despesas_qs.aggregate(
        valor=Sum('valor'),
        quantidade=Count('id'),
    )
    despesas_qs = despesas_qs.order_by('-data', '-id')
    paginator = Paginator(despesas_qs, 50)
    page_obj = paginator.get_page(request.GET.get('page') or 1)
    query_params = request.GET.copy()
    query_params.pop('page', None)

    categorias = CategoriaDespesaSalao.objects.filter(ativo=True).order_by('nome')
    subcategorias = SubcategoriaDespesaSalao.objects.filter(ativo=True).select_related('categoria').order_by(
        'categoria__nome', 'nome'
    )
    if categoria_id:
        subcategorias = subcategorias.filter(categoria_id=categoria_id)

    context = {
        'active_tab': 'dashboard',
        'ano': ano,
        'mes': mes,
        'month_options': MONTH_OPTIONS,
        'year_options': _build_year_options(),
        'categorias_ativas': categorias,
        'subcategorias_ativas': subcategorias,
        'filtro_categoria_id': categoria_id,
        'filtro_subcategoria_id': subcategoria_id,
        'page_obj': page_obj,
        'pagination_query': query_params.urlencode(),
        'totais': {
            'valor': resumo_totais['valor'] or Decimal('0.00'),
            'quantidade': resumo_totais['quantidade'] or 0,
        },
    }
    return render(request, 'salao/grid_despesas.html', context)

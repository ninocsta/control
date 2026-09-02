from io import BytesIO
from datetime import date
from decimal import Decimal, ROUND_HALF_UP

from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from .models import (
    CompraEstoqueItemSalao,
    CompraEstoqueSalao,
    CategoriaDespesaSalao,
    ComissaoMensalSalao,
    DespesaSalao,
    FormaPagamentoSalao,
    LancamentoSalao,
    MovimentoEstoqueSalao,
    ProdutoSalao,
    ServicoSalao,
    SubcategoriaDespesaSalao,
    TaxaFormaPagamentoSalao,
    TransacaoIgnoradaSalao,
)
from .views import (
    _calcular_bruto_com_taxa_repassada,
    _calcular_liquido_com_taxa,
    _parse_extrato_csv,
)


class SalaoViewsTests(TestCase):
    def setUp(self):
        User = get_user_model()
        self.superuser = User.objects.create_superuser(
            username='admin',
            email='admin@example.com',
            password='test1234',
        )

        self.servico = ServicoSalao.objects.create(
            codigo='C01',
            nome='Corte Feminino',
            valor_padrao=Decimal('80.00'),
            ativo=True,
        )
        self.categoria = CategoriaDespesaSalao.objects.create(
            nome='Produtos',
            ativo=True,
        )
        self.subcategoria = SubcategoriaDespesaSalao.objects.create(
            categoria=self.categoria,
            nome='Auxiliar Maria',
            ativo=True,
        )
        self.produto = ProdutoSalao.objects.create(
            codigo='P01',
            nome='Máscara Capilar',
            unidade='UN',
            valor_venda_padrao=Decimal('35.00'),
            estoque_minimo=Decimal('2.000'),
            saldo_atual=Decimal('0.000'),
            custo_medio_atual=Decimal('0.00'),
            ativo=True,
        )

        self.forma_nao_informado, _ = FormaPagamentoSalao.objects.get_or_create(
            codigo='0',
            defaults={
                'nome': 'Não informado',
                'ativo': True,
                'aceita_parcelamento': False,
            },
        )
        self.forma_pix, _ = FormaPagamentoSalao.objects.get_or_create(
            codigo='1',
            defaults={
                'nome': 'PIX',
                'ativo': True,
                'aceita_parcelamento': False,
            },
        )
        self.forma_credito, _ = FormaPagamentoSalao.objects.get_or_create(
            codigo='2',
            defaults={
                'nome': 'Crédito',
                'ativo': True,
                'aceita_parcelamento': True,
            },
        )
        self.forma_dinheiro, _ = FormaPagamentoSalao.objects.get_or_create(
            codigo='3',
            defaults={
                'nome': 'Dinheiro',
                'ativo': True,
                'aceita_parcelamento': False,
            },
        )
        self.forma_debito, _ = FormaPagamentoSalao.objects.get_or_create(
            codigo='4',
            defaults={
                'nome': 'Débito',
                'ativo': True,
                'aceita_parcelamento': False,
            },
        )

        TaxaFormaPagamentoSalao.objects.update_or_create(
            forma_pagamento=self.forma_nao_informado,
            parcelas=1,
            defaults={'percentual': Decimal('0.00')},
        )
        TaxaFormaPagamentoSalao.objects.update_or_create(
            forma_pagamento=self.forma_pix,
            parcelas=1,
            defaults={'percentual': Decimal('0.00')},
        )
        TaxaFormaPagamentoSalao.objects.update_or_create(
            forma_pagamento=self.forma_dinheiro,
            parcelas=1,
            defaults={'percentual': Decimal('0.00')},
        )
        TaxaFormaPagamentoSalao.objects.update_or_create(
            forma_pagamento=self.forma_debito,
            parcelas=1,
            defaults={'percentual': Decimal('3.00')},
        )
        TaxaFormaPagamentoSalao.objects.update_or_create(
            forma_pagamento=self.forma_credito,
            parcelas=1,
            defaults={'percentual': Decimal('4.00')},
        )
        TaxaFormaPagamentoSalao.objects.update_or_create(
            forma_pagamento=self.forma_credito,
            parcelas=2,
            defaults={'percentual': Decimal('5.00')},
        )

    def _login(self):
        self.client.force_login(self.superuser)

    def _create_lancamento(
        self,
        *,
        data,
        valor_bruto,
        forma_pagamento=None,
        parcelas=1,
        taxa_percentual=Decimal('0.00'),
        permuta=False,
        sem_comissao=False,
    ):
        if permuta:
            forma = self.forma_nao_informado
            parcelas = 1
            taxa_percentual = Decimal('0.00')
            valor_taxa = Decimal('0.00')
            valor_liquido = valor_bruto.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        else:
            forma = forma_pagamento or self.forma_dinheiro
            valor_taxa = (valor_bruto * taxa_percentual / Decimal('100.00')).quantize(
                Decimal('0.01'),
                rounding=ROUND_HALF_UP,
            )
            valor_liquido = (valor_bruto - valor_taxa).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        return LancamentoSalao.objects.create(
            data=data,
            servico=self.servico,
            forma_pagamento=forma,
            parcelas=parcelas,
            permuta=permuta,
            sem_comissao=sem_comissao,
            valor_bruto=valor_bruto,
            taxa_percentual_aplicada=taxa_percentual,
            valor_taxa=valor_taxa,
            valor_cobrado=valor_liquido,
        )

    def test_dashboard_auto_create_comissao(self):
        self._login()
        self.assertEqual(ComissaoMensalSalao.objects.count(), 0)

        response = self.client.get(reverse('salao:dashboard'), {'ano': 2026, 'mes': 2})

        self.assertEqual(response.status_code, 200)
        comissao = ComissaoMensalSalao.objects.get(ano=2026, mes=2)
        self.assertEqual(comissao.percentual, Decimal('20.00'))

    def test_dashboard_calculo_sem_override(self):
        self._login()

        self._create_lancamento(
            data=date(2026, 3, 10),
            valor_bruto=Decimal('200.00'),
            forma_pagamento=self.forma_dinheiro,
            taxa_percentual=Decimal('0.00'),
        )
        self._create_lancamento(
            data=date(2026, 3, 12),
            valor_bruto=Decimal('100.00'),
            forma_pagamento=self.forma_dinheiro,
            taxa_percentual=Decimal('0.00'),
        )
        DespesaSalao.objects.create(
            data=date(2026, 3, 13),
            categoria=self.categoria,
            valor=Decimal('50.00'),
            observacao='Compra mensal',
        )

        response = self.client.get(reverse('salao:dashboard'), {'ano': 2026, 'mes': 3})
        self.assertEqual(response.status_code, 200)

        self.assertEqual(response.context['faturamento_bruto'], Decimal('300.00'))
        self.assertEqual(response.context['taxas_total'], Decimal('0.00'))
        self.assertEqual(response.context['comissao_calculada'], Decimal('60.00'))
        self.assertEqual(response.context['despesas_total'], Decimal('50.00'))
        self.assertEqual(response.context['lucro'], Decimal('190.00'))
        self.assertEqual(response.context['permuta_total_mes'], Decimal('0.00'))
        self.assertIn('meta_bullet_chart', response.context)

    def test_grid_totais_normais_batem_com_o_dashboard(self):
        self._login()

        self._create_lancamento(
            data=date(2026, 3, 10),
            valor_bruto=Decimal('200.00'),
            forma_pagamento=self.forma_dinheiro,
        )
        self._create_lancamento(
            data=date(2026, 3, 11),
            valor_bruto=Decimal('100.00'),
            permuta=True,
        )
        self._create_lancamento(
            data=date(2026, 3, 12),
            valor_bruto=Decimal('50.00'),
            forma_pagamento=self.forma_dinheiro,
            sem_comissao=True,
        )

        grid = self.client.get(reverse('salao:grid_lancamentos'), {'ano': 2026, 'mes': 3})
        dashboard = self.client.get(reverse('salao:dashboard'), {'ano': 2026, 'mes': 3})
        totais = grid.context['totais']

        # Total geral soma tudo; o subtotal "normais" é a base do dashboard.
        self.assertEqual(totais['bruto'], Decimal('350.00'))
        self.assertEqual(totais['quantidade'], 3)
        self.assertEqual(totais['liquido_normais'], dashboard.context['faturamento_bruto'])
        self.assertEqual(totais['bruto_normais'], dashboard.context['faturamento_bruto_cliente'])
        self.assertEqual(totais['quantidade_normais'], 1)

    # ------------------------------------------------------------------
    # Conferência bancária
    # ------------------------------------------------------------------
    CSV_EXTRATO = (
        'Data e hora,Meio - Meio,Meio - Bandeira,Meio - Parcelas,Tipo - Origem,'
        'Tipo - Dados adicionais,Identificador,Status,Valor (R$),Líquido (R$),'
        'Taxa Aplicada - Valor(R$),Taxa Aplicada - Aplicada(%),Plano,NSU,Origem - Nome\n'
        # taxa embutida no cliente: o líquido é que fecha com o valor do serviço
        '15/03/2026 14:02,Crédito,visa,À Vista,Maquininha,\'-,TX001,Aprovada,'
        '"165,21","160,01","\'- 5,19",3.14,1 Dia Útil,S1,""\n'
        # taxa paga pelo salão: o bruto é que fecha
        '15/03/2026 15:10,Pix,Pix,À Vista,Maquininha,\'-,TX002,Aprovada,'
        '"170,00","170,00","0,00",0,Outro,S2,""\n'
        # recebimento por fora, sem lançamento no salão
        '15/03/2026 12:06,Pix,Pix,À Vista,Conta Inteligente,\'-,TX003,Aprovada,'
        '"8.000,00","8.000,00","0,00",0,Outro,S3,JEAN\n'
        # cancelada: não entra na conferência
        '15/03/2026 16:00,Crédito,visa,À Vista,Maquininha,\'-,TX004,Cancelada,'
        '"90,00","87,00","\'- 3,00",3.14,1 Dia Útil,S4,""\n'
        # outra competência: descartada na importação
        '15/04/2026 10:00,Pix,Pix,À Vista,Maquininha,\'-,TX005,Aprovada,'
        '"50,00","50,00","0,00",0,Outro,S5,""\n'
    )

    def _importar_extrato(self, csv_texto=None):
        arquivo = SimpleUploadedFile(
            'extrato.csv',
            (csv_texto or self.CSV_EXTRATO).encode('utf-8'),
            content_type='text/csv',
        )
        return self.client.post(
            reverse('salao:conferencia'),
            {'action': 'importar_extrato', 'ano': 2026, 'mes': 3, 'extrato': arquivo},
            follow=True,
        )

    def _lancamentos_do_extrato(self):
        """Os dois lançamentos que correspondem a TX001 (líquido) e TX002 (bruto)."""
        self._create_lancamento(
            data=date(2026, 3, 15),
            valor_bruto=Decimal('160.00'),
            forma_pagamento=self.forma_credito,
            taxa_percentual=Decimal('2.00'),
        )
        self._create_lancamento(
            data=date(2026, 3, 15),
            valor_bruto=Decimal('170.00'),
            forma_pagamento=self.forma_pix,
        )

    def test_conferencia_parser_le_valores_pt_br_e_percentual_com_ponto(self):
        transacoes = _parse_extrato_csv(BytesIO(self.CSV_EXTRATO.encode('utf-8')))
        # o parser não descarta nada: a linha de abril entra com a própria data
        self.assertEqual(len(transacoes), 5)
        por_id = {t['identificador']: t for t in transacoes}
        self.assertEqual((por_id['TX005']['ano'], por_id['TX005']['mes']), (2026, 4))
        self.assertEqual(por_id['TX001']['bruto'], '165.21')
        self.assertEqual(por_id['TX001']['liquido'], '160.01')
        self.assertEqual(por_id['TX001']['taxa_percentual'], '3.14')
        self.assertEqual(por_id['TX003']['bruto'], '8000.00')
        self.assertEqual(por_id['TX004']['status'], 'Cancelada')

    def test_conferencia_pareia_pelo_liquido_quando_a_taxa_foi_embutida(self):
        self._login()
        # Serviço de R$ 160,00 cobrado como R$ 165,21 no cartão para o líquido fechar em 160.
        lancamento = self._create_lancamento(
            data=date(2026, 3, 15),
            valor_bruto=Decimal('160.00'),
            forma_pagamento=self.forma_credito,
            taxa_percentual=Decimal('2.00'),
        )
        self._importar_extrato()

        response = self.client.get(reverse('salao:conferencia'), {'ano': 2026, 'mes': 3})
        dia = response.context['dias'][0]
        par = next(p for p in dia['pares'] if p['lancamentos'][0].id == lancamento.id)
        self.assertEqual(par['criterio'], 'liquido')
        self.assertEqual(par['transacao']['identificador'], 'TX001')
        # taxa real 3,14% contra 2,00% cadastrado
        self.assertEqual(par['diferenca_taxa'], Decimal('1.14'))

    def test_conferencia_pareia_pelo_bruto_quando_o_salao_pagou_a_taxa(self):
        self._login()
        lancamento = self._create_lancamento(
            data=date(2026, 3, 15),
            valor_bruto=Decimal('170.00'),
            forma_pagamento=self.forma_pix,
            taxa_percentual=Decimal('0.00'),
        )
        self._importar_extrato()

        response = self.client.get(reverse('salao:conferencia'), {'ano': 2026, 'mes': 3})
        par = next(
            p for p in response.context['dias'][0]['pares']
            if p['lancamentos'][0].id == lancamento.id
        )
        self.assertEqual(par['criterio'], 'bruto')
        self.assertEqual(par['transacao']['identificador'], 'TX002')
        self.assertEqual(par['diferenca_valor'], Decimal('0.00'))

    def test_conferencia_nao_pareia_meios_de_pagamento_diferentes(self):
        """Dinheiro de R$ 100,00 não pode casar com Crédito de R$ 103,15/99,91."""
        self._login()
        csv_credito = (
            'Data e hora,Meio - Meio,Meio - Bandeira,Meio - Parcelas,Tipo - Origem,'
            'Tipo - Dados adicionais,Identificador,Status,Valor (R$),Líquido (R$),'
            'Taxa Aplicada - Valor(R$),Taxa Aplicada - Aplicada(%),Plano,NSU,Origem - Nome\n'
            '15/03/2026 16:06,Crédito,mastercard,À Vista,Maquininha,\'-,TXC,Aprovada,'
            '"103,15","99,91","\'- 3,24",3.14,1 Dia Útil,S9,""\n'
        )
        lancamento = self._create_lancamento(
            data=date(2026, 3, 15),
            valor_bruto=Decimal('100.00'),
            forma_pagamento=self.forma_dinheiro,
        )
        self._importar_extrato(csv_credito)

        response = self.client.get(reverse('salao:conferencia'), {'ano': 2026, 'mes': 3})
        dia = response.context['dias'][0]
        self.assertEqual(dia['pares'], [])
        self.assertEqual([l.id for l in dia['sem_par_sistema']], [lancamento.id])
        self.assertEqual([t['identificador'] for t in dia['sem_par_extrato']], ['TXC'])

    def test_conferencia_tolerancia_so_vale_dentro_do_mesmo_meio(self):
        self._login()
        # mesmo meio e centavo de diferença: casa
        credito = self._create_lancamento(
            data=date(2026, 3, 15),
            valor_bruto=Decimal('160.00'),
            forma_pagamento=self.forma_credito,
            taxa_percentual=Decimal('3.15'),
        )
        self._importar_extrato()

        response = self.client.get(reverse('salao:conferencia'), {'ano': 2026, 'mes': 3})
        par = next(
            p for p in response.context['dias'][0]['pares']
            if p['lancamentos'][0].id == credito.id
        )
        self.assertEqual(par['transacao']['identificador'], 'TX001')
        self.assertEqual(par['diferenca_valor'], Decimal('-0.01'))

    def test_conferencia_importa_extrato_com_o_filtro_em_outra_competencia(self):
        """O CSV é de 03/2026; importar com a tela em 08/2028 não pode perder nada."""
        self._login()
        self._lancamentos_do_extrato()

        response = self.client.post(
            reverse('salao:conferencia'),
            {
                'action': 'importar_extrato',
                'ano': 2028,
                'mes': 8,
                'extrato': SimpleUploadedFile(
                    'extrato.csv', self.CSV_EXTRATO.encode('utf-8'), content_type='text/csv',
                ),
            },
            follow=True,
        )
        # redireciona para a competência do próprio arquivo
        self.assertEqual((response.context['ano'], response.context['mes']), (2026, 3))
        self.assertTrue(response.context['tem_extrato'])
        self.assertEqual(len(response.context['dias'][0]['pares']), 2)

    def test_conferencia_avisa_quando_o_extrato_e_de_outro_mes(self):
        self._login()
        self._importar_extrato()

        response = self.client.get(reverse('salao:conferencia'), {'ano': 2026, 'mes': 12})
        self.assertFalse(response.context['tem_extrato'])
        cobertura = {
            (item['ano'], item['mes']) for item in response.context['extrato_outra_competencia']
        }
        self.assertEqual(cobertura, {(2026, 3), (2026, 4)})

    def test_conferencia_extrato_sobrevive_a_troca_de_competencia(self):
        self._login()
        self._lancamentos_do_extrato()
        self._importar_extrato()

        # passa por abril (só TX005) e volta para março sem reimportar
        abril = self.client.get(reverse('salao:conferencia'), {'ano': 2026, 'mes': 4})
        self.assertTrue(abril.context['tem_extrato'])
        marco = self.client.get(reverse('salao:conferencia'), {'ano': 2026, 'mes': 3})
        self.assertEqual(len(marco.context['dias'][0]['pares']), 2)

    def _create_venda_produto(self, *, data, valor_bruto, forma_pagamento=None, taxa=Decimal('0.00')):
        valor_taxa = (valor_bruto * taxa / Decimal('100.00')).quantize(
            Decimal('0.01'), rounding=ROUND_HALF_UP,
        )
        return MovimentoEstoqueSalao.objects.create(
            data=data,
            produto=self.produto,
            tipo=MovimentoEstoqueSalao.TIPO_SAIDA,
            motivo=MovimentoEstoqueSalao.MOTIVO_VENDA,
            quantidade=Decimal('1.000'),
            valor_bruto_venda=valor_bruto,
            taxa_percentual_aplicada=taxa,
            valor_taxa=valor_taxa,
            valor_liquido_venda=valor_bruto - valor_taxa,
            forma_pagamento=forma_pagamento or self.forma_pix,
        )

    def test_conferencia_inclui_venda_de_produto(self):
        """Os R$ 105,00 do extrato eram venda de produto, não serviço."""
        self._login()
        csv_105 = (
            'Data e hora,Meio - Meio,Meio - Bandeira,Meio - Parcelas,Tipo - Origem,'
            'Tipo - Dados adicionais,Identificador,Status,Valor (R$),Líquido (R$),'
            'Taxa Aplicada - Valor(R$),Taxa Aplicada - Aplicada(%),Plano,NSU,Origem - Nome\n'
            '15/03/2026 10:40,Pix,Pix,À Vista,Conta Inteligente,\'-,TX105,Aprovada,'
            '"105,00","105,00","0,00",0,Outro,S11,NEIVA\n'
        )
        venda = self._create_venda_produto(data=date(2026, 3, 15), valor_bruto=Decimal('105.00'))
        self._importar_extrato(csv_105)

        response = self.client.get(reverse('salao:conferencia'), {'ano': 2026, 'mes': 3})
        dia = response.context['dias'][0]
        self.assertEqual(len(dia['pares']), 1)
        item = dia['pares'][0]['lancamentos'][0]
        self.assertEqual(item.tipo, 'produto')
        self.assertEqual(item.chave, f'produto:{venda.id}')
        self.assertFalse(item.ajustavel)

    def test_conferencia_soma_servico_com_produto_na_mesma_transacao(self):
        """R$ 564,56 = progressiva de R$ 383,47 mais produtos."""
        self._login()
        csv_564 = (
            'Data e hora,Meio - Meio,Meio - Bandeira,Meio - Parcelas,Tipo - Origem,'
            'Tipo - Dados adicionais,Identificador,Status,Valor (R$),Líquido (R$),'
            'Taxa Aplicada - Valor(R$),Taxa Aplicada - Aplicada(%),Plano,NSU,Origem - Nome\n'
            '15/03/2026 17:58,Crédito,visa,3,Maquininha,\'-,TX564,Aprovada,'
            '"564,56","524,00","\'- 40,56",7.19,1 Dia Útil,S12,GARSKE\n'
        )
        servico = self._create_lancamento(
            data=date(2026, 3, 15),
            valor_bruto=Decimal('383.47'),
            forma_pagamento=self.forma_credito,
            taxa_percentual=Decimal('6.12'),
        )
        venda = self._create_venda_produto(
            data=date(2026, 3, 15),
            valor_bruto=Decimal('181.09'),
            forma_pagamento=self.forma_credito,
        )
        self._importar_extrato(csv_564)

        response = self.client.get(reverse('salao:conferencia'), {'ano': 2026, 'mes': 3})
        par = response.context['dias'][0]['pares'][0]
        self.assertTrue(par['combinado'])
        self.assertEqual(
            {i.chave for i in par['lancamentos']},
            {f'servico:{servico.id}', f'produto:{venda.id}'},
        )
        self.assertEqual(par['total_lancado'], Decimal('564.56'))

    def test_conferencia_concilia_pagamento_que_cobre_dias_diferentes(self):
        """Pix de R$ 200,00 no dia 07 cobrindo atendimentos do dia 01 e do dia 07."""
        self._login()
        csv_200 = (
            'Data e hora,Meio - Meio,Meio - Bandeira,Meio - Parcelas,Tipo - Origem,'
            'Tipo - Dados adicionais,Identificador,Status,Valor (R$),Líquido (R$),'
            'Taxa Aplicada - Valor(R$),Taxa Aplicada - Aplicada(%),Plano,NSU,Origem - Nome\n'
            '07/03/2026 07:16,Pix,Pix,À Vista,Maquininha,\'-,TX200,Aprovada,'
            '"200,00","200,00","0,00",0,Outro,S13,Ana Luisa\n'
        )
        dia01 = self._create_lancamento(
            data=date(2026, 3, 1), valor_bruto=Decimal('100.00'), forma_pagamento=self.forma_pix,
        )
        dia07 = self._create_lancamento(
            data=date(2026, 3, 7), valor_bruto=Decimal('100.00'), forma_pagamento=self.forma_pix,
        )
        self._importar_extrato(csv_200)

        response = self.client.get(reverse('salao:conferencia'), {'ano': 2026, 'mes': 3})
        dias = {d['dia']: d for d in response.context['dias']}
        # o par aparece no dia em que o dinheiro entrou
        par = dias[7]['pares'][0]
        self.assertTrue(par['entre_dias'])
        self.assertEqual(
            {i.id for i in par['lancamentos']}, {dia01.id, dia07.id},
        )
        self.assertEqual(dias[1]['sem_par_sistema'], [])
        self.assertEqual(dias[7]['sem_par_extrato'], [])

    def test_conferencia_nao_cruza_dias_alem_da_janela(self):
        self._login()
        csv_200 = (
            'Data e hora,Meio - Meio,Meio - Bandeira,Meio - Parcelas,Tipo - Origem,'
            'Tipo - Dados adicionais,Identificador,Status,Valor (R$),Líquido (R$),'
            'Taxa Aplicada - Valor(R$),Taxa Aplicada - Aplicada(%),Plano,NSU,Origem - Nome\n'
            '25/03/2026 07:16,Pix,Pix,À Vista,Maquininha,\'-,TX200,Aprovada,'
            '"200,00","200,00","0,00",0,Outro,S13,Ana\n'
        )
        self._create_lancamento(
            data=date(2026, 3, 1), valor_bruto=Decimal('100.00'), forma_pagamento=self.forma_pix,
        )
        self._create_lancamento(
            data=date(2026, 3, 25), valor_bruto=Decimal('100.00'), forma_pagamento=self.forma_pix,
        )
        self._importar_extrato(csv_200)

        response = self.client.get(reverse('salao:conferencia'), {'ano': 2026, 'mes': 3})
        dias = {d['dia']: d for d in response.context['dias']}
        self.assertEqual(dias[25]['pares'], [])
        self.assertEqual(len(dias[25]['sem_par_extrato']), 1)

    def test_conferencia_marca_venda_de_produto_como_conferida(self):
        self._login()
        venda = self._create_venda_produto(data=date(2026, 3, 15), valor_bruto=Decimal('105.00'))

        response = self.client.post(
            reverse('salao:conferencia'),
            {'action': 'toggle_conferido', 'ano': 2026, 'mes': 3,
             'item': f'produto:{venda.id}', 'conferido': 'on'},
            HTTP_X_REQUESTED_WITH='XMLHttpRequest',
        )
        self.assertEqual(response.json()['conferido'], True)
        venda.refresh_from_db()
        self.assertTrue(venda.conferido)
        self.assertIsNotNone(venda.conferido_em)

    def test_conferencia_conferir_dia_pega_servico_e_produto(self):
        self._login()
        self._create_lancamento(
            data=date(2026, 3, 15), valor_bruto=Decimal('80.00'), forma_pagamento=self.forma_pix,
        )
        venda = self._create_venda_produto(data=date(2026, 3, 15), valor_bruto=Decimal('105.00'))

        self.client.post(
            reverse('salao:conferencia'),
            {'action': 'conferir_dia', 'ano': 2026, 'mes': 3, 'dia': 15},
        )
        venda.refresh_from_db()
        self.assertTrue(venda.conferido)
        self.assertEqual(
            LancamentoSalao.objects.filter(data=date(2026, 3, 15), conferido=True).count(), 1,
        )

    def test_conferencia_conferir_dia_nao_liga_filtro_de_dia(self):
        """Conferir/desmarcar um dia mantém a tela em que o usuário estava."""
        self._login()
        self._create_lancamento(
            data=date(2026, 3, 15), valor_bruto=Decimal('80.00'), forma_pagamento=self.forma_pix,
        )

        for acao in ('conferir_dia', 'desconferir_dia'):
            with self.subTest(acao=acao):
                response = self.client.post(
                    reverse('salao:conferencia'),
                    {
                        'action': acao,
                        'ano': 2026, 'mes': 3, 'dia': 15,
                        'retorno': 'ano=2026&mes=3',
                    },
                )
                self.assertEqual(response.status_code, 302)
                self.assertNotIn('dia=', response['Location'])

    def test_conferencia_conferir_dia_preserva_o_filtro_que_o_usuario_pediu(self):
        self._login()
        self._create_lancamento(
            data=date(2026, 3, 1), valor_bruto=Decimal('80.00'), forma_pagamento=self.forma_pix,
        )

        response = self.client.post(
            reverse('salao:conferencia'),
            {
                'action': 'conferir_dia',
                'ano': 2026, 'mes': 3, 'dia': 1,
                'retorno': 'ano=2026&mes=3&dia=1&pendentes=1',
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertIn('dia=1', response['Location'])
        self.assertIn('pendentes=1', response['Location'])

    def test_conferencia_ajustar_lancamento_nao_liga_filtro_de_dia(self):
        self._login()
        lancamento = self._create_lancamento(
            data=date(2026, 3, 15), valor_bruto=Decimal('100.00'), forma_pagamento=self.forma_pix,
        )

        response = self.client.post(
            reverse('salao:conferencia'),
            {
                'action': 'ajustar_lancamento',
                'ano': 2026, 'mes': 3,
                'item': f'servico:{lancamento.id}',
                'valor_bruto': '105,00',
                'forma_pagamento_id': self.forma_pix.id,
                'parcelas': 1,
                'retorno': 'ano=2026&mes=3',
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertNotIn('dia=', response['Location'])

    def test_conferencia_ignorar_transacao_volta_para_a_mesma_tela(self):
        self._login()
        self._importar_extrato()

        response = self.client.post(
            reverse('salao:conferencia'),
            {
                'action': 'ignorar_transacao',
                'ano': 2026, 'mes': 3,
                'identificador': 'TX003',
                'retorno': 'ano=2026&mes=3',
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertNotIn('dia=', response['Location'])
        self.assertIn('ano=2026&mes=3', response['Location'])

    def test_conferencia_soma_lancamentos_para_fechar_uma_transacao(self):
        """Cliente paga R$ 290,00 numa passada só, cobrindo dois serviços."""
        self._login()
        csv_290 = (
            'Data e hora,Meio - Meio,Meio - Bandeira,Meio - Parcelas,Tipo - Origem,'
            'Tipo - Dados adicionais,Identificador,Status,Valor (R$),Líquido (R$),'
            'Taxa Aplicada - Valor(R$),Taxa Aplicada - Aplicada(%),Plano,NSU,Origem - Nome\n'
            '15/03/2026 11:00,Pix,Pix,À Vista,Maquininha,\'-,TX290,Aprovada,'
            '"290,00","290,00","0,00",0,Outro,S10,""\n'
        )
        a = self._create_lancamento(
            data=date(2026, 3, 15), valor_bruto=Decimal('170.00'), forma_pagamento=self.forma_pix,
        )
        b = self._create_lancamento(
            data=date(2026, 3, 15), valor_bruto=Decimal('120.00'), forma_pagamento=self.forma_pix,
        )
        self._importar_extrato(csv_290)

        response = self.client.get(reverse('salao:conferencia'), {'ano': 2026, 'mes': 3})
        dia = response.context['dias'][0]
        self.assertEqual(len(dia['pares']), 1)
        par = dia['pares'][0]
        self.assertTrue(par['combinado'])
        self.assertEqual({l.id for l in par['lancamentos']}, {a.id, b.id})
        self.assertEqual(par['total_lancado'], Decimal('290.00'))
        self.assertEqual(dia['sem_par_sistema'], [])
        self.assertEqual(dia['sem_par_extrato'], [])

    def test_conferencia_nao_soma_lancamentos_de_meios_diferentes(self):
        self._login()
        csv_290 = (
            'Data e hora,Meio - Meio,Meio - Bandeira,Meio - Parcelas,Tipo - Origem,'
            'Tipo - Dados adicionais,Identificador,Status,Valor (R$),Líquido (R$),'
            'Taxa Aplicada - Valor(R$),Taxa Aplicada - Aplicada(%),Plano,NSU,Origem - Nome\n'
            '15/03/2026 11:00,Pix,Pix,À Vista,Maquininha,\'-,TX290,Aprovada,'
            '"290,00","290,00","0,00",0,Outro,S10,""\n'
        )
        self._create_lancamento(
            data=date(2026, 3, 15), valor_bruto=Decimal('170.00'), forma_pagamento=self.forma_pix,
        )
        self._create_lancamento(
            data=date(2026, 3, 15), valor_bruto=Decimal('120.00'), forma_pagamento=self.forma_dinheiro,
        )
        self._importar_extrato(csv_290)

        response = self.client.get(reverse('salao:conferencia'), {'ano': 2026, 'mes': 3})
        dia = response.context['dias'][0]
        self.assertEqual(dia['pares'], [])
        self.assertEqual(len(dia['sem_par_sistema']), 2)

    def test_conferencia_ajusta_lancamento_e_recalcula_a_taxa(self):
        """Lançado R$ 100,00 mas o extrato mostra R$ 105,00: corrigir na tela."""
        self._login()
        TaxaFormaPagamentoSalao.objects.update_or_create(
            forma_pagamento=self.forma_credito,
            parcelas=1,
            defaults={'percentual': Decimal('3.15')},
        )
        lancamento = self._create_lancamento(
            data=date(2026, 3, 15),
            valor_bruto=Decimal('100.00'),
            forma_pagamento=self.forma_pix,
        )

        self.client.post(
            reverse('salao:conferencia'),
            {
                'action': 'ajustar_lancamento',
                'ano': 2026,
                'mes': 3,
                'item': f'servico:{lancamento.id}',
                'valor_bruto': '105,00',
                'forma_pagamento_id': self.forma_credito.id,
                'parcelas': 1,
            },
        )

        lancamento.refresh_from_db()
        self.assertEqual(lancamento.valor_bruto, Decimal('105.00'))
        self.assertEqual(lancamento.forma_pagamento_id, self.forma_credito.id)
        self.assertEqual(lancamento.taxa_percentual_aplicada, Decimal('3.15'))
        self.assertEqual(lancamento.valor_taxa, Decimal('3.31'))
        self.assertEqual(lancamento.valor_cobrado, Decimal('101.69'))

    def test_conferencia_ajusta_venda_de_produto_sem_mexer_no_estoque(self):
        """O erro pode estar no produto; ajustar o preço não pode mexer no saldo."""
        self._login()
        TaxaFormaPagamentoSalao.objects.update_or_create(
            forma_pagamento=self.forma_credito, parcelas=3,
            defaults={'percentual': Decimal('7.19')},
        )
        venda = self._create_venda_produto(
            data=date(2026, 3, 5),
            valor_bruto=Decimal('89.00'),
            forma_pagamento=self.forma_credito,
        )
        venda.valor_custo_total = Decimal('30.00')
        venda.quantidade = Decimal('2.000')
        venda.save()
        saldo_antes = ProdutoSalao.objects.get(pk=self.produto.pk).saldo_atual

        self.client.post(
            reverse('salao:conferencia'),
            {
                'action': 'ajustar_lancamento',
                'ano': 2026, 'mes': 3,
                'item': f'produto:{venda.id}',
                'valor_bruto': '107,09',
                'forma_pagamento_id': self.forma_credito.id,
                'parcelas': 3,
                'retorno': 'ano=2026&mes=3',
            },
        )

        venda.refresh_from_db()
        self.assertEqual(venda.valor_bruto_venda, Decimal('107.09'))
        self.assertEqual(venda.taxa_percentual_aplicada, Decimal('7.19'))
        self.assertEqual(venda.valor_taxa, Decimal('7.70'))
        self.assertEqual(venda.valor_liquido_venda, Decimal('99.39'))
        # unitário acompanha o novo preço e o lucro é recalculado sobre o custo
        self.assertEqual(venda.valor_venda_unitario, Decimal('53.55'))
        self.assertEqual(venda.lucro_produto, Decimal('69.39'))
        # quantidade e saldo de estoque intactos
        self.assertEqual(venda.quantidade, Decimal('2.000'))
        self.assertEqual(
            ProdutoSalao.objects.get(pk=self.produto.pk).saldo_atual, saldo_antes,
        )

    def test_conferencia_mostra_a_diferenca_das_sobras_do_dia(self):
        """89 + 74 + 383,47 = 546,47 contra uma transação de 564,56."""
        self._login()
        csv_564 = (
            'Data e hora,Meio - Meio,Meio - Bandeira,Meio - Parcelas,Tipo - Origem,'
            'Tipo - Dados adicionais,Identificador,Status,Valor (R$),Líquido (R$),'
            'Taxa Aplicada - Valor(R$),Taxa Aplicada - Aplicada(%),Plano,NSU,Origem - Nome\n'
            '05/03/2026 17:58,Crédito,elo,3,Maquininha,\'-,TX564,Aprovada,'
            '"564,56","523,92","\'- 40,64",7.19,1 Dia Útil,S12,GARSKE\n'
        )
        self._create_lancamento(
            data=date(2026, 3, 5), valor_bruto=Decimal('383.47'),
            forma_pagamento=self.forma_credito, taxa_percentual=Decimal('6.12'),
        )
        for valor in ('89.00', '74.00'):
            self._create_venda_produto(
                data=date(2026, 3, 5), valor_bruto=Decimal(valor),
                forma_pagamento=self.forma_credito,
            )
        self._importar_extrato(csv_564)

        response = self.client.get(reverse('salao:conferencia'), {'ano': 2026, 'mes': 3})
        sobras = response.context['dias'][0]['sobras']
        self.assertEqual(sobras['soma_sistema'], Decimal('546.47'))
        self.assertEqual(sobras['soma_extrato'], Decimal('564.56'))
        self.assertEqual(sobras['diferenca'], Decimal('-18.09'))

    def test_conferencia_sem_sobras_dos_dois_lados_nao_mostra_resumo(self):
        self._login()
        self._create_lancamento(
            data=date(2026, 3, 5), valor_bruto=Decimal('80.00'), forma_pagamento=self.forma_pix,
        )
        response = self.client.get(reverse('salao:conferencia'), {'ano': 2026, 'mes': 3})
        self.assertIsNone(response.context['dias'][0]['sobras'])

    def test_conferencia_ajuste_recusa_forma_sem_taxa_cadastrada(self):
        self._login()
        lancamento = self._create_lancamento(
            data=date(2026, 3, 15),
            valor_bruto=Decimal('100.00'),
            forma_pagamento=self.forma_pix,
        )
        TaxaFormaPagamentoSalao.objects.filter(
            forma_pagamento=self.forma_credito, parcelas=7,
        ).delete()

        self.client.post(
            reverse('salao:conferencia'),
            {
                'action': 'ajustar_lancamento',
                'ano': 2026, 'mes': 3,
                'item': f'servico:{lancamento.id}',
                'valor_bruto': '105,00',
                'forma_pagamento_id': self.forma_credito.id,
                'parcelas': 7,
            },
        )
        lancamento.refresh_from_db()
        self.assertEqual(lancamento.valor_bruto, Decimal('100.00'))

    def test_conferencia_lista_transacao_do_extrato_sem_lancamento(self):
        self._login()
        self._lancamentos_do_extrato()
        self._importar_extrato()

        response = self.client.get(reverse('salao:conferencia'), {'ano': 2026, 'mes': 3})
        sem_par = response.context['dias'][0]['sem_par_extrato']
        self.assertEqual([t['identificador'] for t in sem_par], ['TX003'])
        # cancelada não entra na conciliação
        self.assertEqual(
            [t['identificador'] for t in response.context['transacoes_recusadas']],
            ['TX004'],
        )

    def test_conferencia_transacao_ignorada_some_do_pareamento_e_persiste(self):
        self._login()
        self._lancamentos_do_extrato()
        self._importar_extrato()

        self.client.post(
            reverse('salao:conferencia'),
            {
                'action': 'ignorar_transacao',
                'ano': 2026,
                'mes': 3,
                'identificador': 'TX003',
                'referencia': 'recebimento por fora',
            },
        )
        self.assertTrue(TransacaoIgnoradaSalao.objects.filter(identificador='TX003').exists())

        response = self.client.get(reverse('salao:conferencia'), {'ano': 2026, 'mes': 3})
        self.assertEqual(response.context['dias'][0]['sem_par_extrato'], [])
        self.assertEqual(
            [t['identificador'] for t in response.context['transacoes_ignoradas']],
            ['TX003'],
        )

        # reimportar o mesmo CSV mantém o recebimento por fora de lado
        self._importar_extrato()
        response = self.client.get(reverse('salao:conferencia'), {'ano': 2026, 'mes': 3})
        self.assertEqual(response.context['dias'][0]['sem_par_extrato'], [])

    def test_conferencia_permuta_fica_fora_do_pareamento(self):
        self._login()
        self._create_lancamento(
            data=date(2026, 3, 15),
            valor_bruto=Decimal('170.00'),
            permuta=True,
        )
        self._importar_extrato()

        response = self.client.get(reverse('salao:conferencia'), {'ano': 2026, 'mes': 3})
        dia = response.context['dias'][0]
        self.assertEqual(len(dia['permutas']), 1)
        self.assertEqual(dia['pares'], [])
        # a permuta não consumiu a transação de R$ 170,00 do extrato
        self.assertIn('TX002', [t['identificador'] for t in dia['sem_par_extrato']])

    def test_conferencia_marca_e_desmarca_conferido_via_ajax(self):
        self._login()
        lancamento = self._create_lancamento(
            data=date(2026, 3, 15),
            valor_bruto=Decimal('170.00'),
            forma_pagamento=self.forma_pix,
        )

        marcar = self.client.post(
            reverse('salao:conferencia'),
            {'action': 'toggle_conferido', 'ano': 2026, 'mes': 3,
             'item': f'servico:{lancamento.id}', 'conferido': 'on'},
            HTTP_X_REQUESTED_WITH='XMLHttpRequest',
        )
        self.assertEqual(marcar.json()['conferido'], True)
        lancamento.refresh_from_db()
        self.assertTrue(lancamento.conferido)
        self.assertIsNotNone(lancamento.conferido_em)

        self.client.post(
            reverse('salao:conferencia'),
            {'action': 'toggle_conferido', 'ano': 2026, 'mes': 3,
             'item': f'servico:{lancamento.id}', 'conferido': ''},
            HTTP_X_REQUESTED_WITH='XMLHttpRequest',
        )
        lancamento.refresh_from_db()
        self.assertFalse(lancamento.conferido)
        self.assertIsNone(lancamento.conferido_em)

    def test_conferencia_conferir_dia_todo_marca_inclusive_permuta(self):
        self._login()
        self._create_lancamento(
            data=date(2026, 3, 15), valor_bruto=Decimal('170.00'), forma_pagamento=self.forma_pix,
        )
        self._create_lancamento(data=date(2026, 3, 15), valor_bruto=Decimal('90.00'), permuta=True)
        fora_do_dia = self._create_lancamento(
            data=date(2026, 3, 16), valor_bruto=Decimal('70.00'), forma_pagamento=self.forma_pix,
        )

        self.client.post(
            reverse('salao:conferencia'),
            {'action': 'conferir_dia', 'ano': 2026, 'mes': 3, 'dia': 15},
        )
        self.assertEqual(
            LancamentoSalao.objects.filter(data=date(2026, 3, 15), conferido=True).count(),
            2,
        )
        fora_do_dia.refresh_from_db()
        self.assertFalse(fora_do_dia.conferido)

    def test_conferencia_filtra_por_dia(self):
        self._login()
        self._create_lancamento(
            data=date(2026, 3, 15), valor_bruto=Decimal('170.00'), forma_pagamento=self.forma_pix,
        )
        self._create_lancamento(
            data=date(2026, 3, 16), valor_bruto=Decimal('70.00'), forma_pagamento=self.forma_pix,
        )

        response = self.client.get(reverse('salao:conferencia'), {'ano': 2026, 'mes': 3, 'dia': 16})
        self.assertEqual([d['dia'] for d in response.context['dias']], [16])

    def test_dashboard_ignora_override_e_usa_comissao_automatica(self):
        self._login()

        self._create_lancamento(
            data=date(2026, 3, 10),
            valor_bruto=Decimal('300.00'),
            forma_pagamento=self.forma_dinheiro,
            taxa_percentual=Decimal('0.00'),
        )
        ComissaoMensalSalao.objects.update_or_create(
            ano=2026,
            mes=3,
            defaults={
                'percentual': Decimal('20.00'),
                'valor_pago_override': Decimal('40.00'),
            },
        )

        response = self.client.get(reverse('salao:dashboard'), {'ano': 2026, 'mes': 3})
        self.assertEqual(response.context['comissao_calculada'], Decimal('60.00'))
        self.assertEqual(response.context['lucro'], Decimal('240.00'))

    def test_dashboard_salva_meta_e_calcula_percentual_atingido(self):
        self._login()

        self._create_lancamento(
            data=date(2026, 3, 10),
            valor_bruto=Decimal('500.00'),
            forma_pagamento=self.forma_dinheiro,
            taxa_percentual=Decimal('0.00'),
        )

        post_response = self.client.post(
            reverse('salao:dashboard'),
            {
                'action': 'update_meta',
                'ano': 2026,
                'mes': 3,
                'meta_faturamento': '1000,00',
            },
        )
        self.assertEqual(post_response.status_code, 302)

        response = self.client.get(reverse('salao:dashboard'), {'ano': 2026, 'mes': 3})
        self.assertEqual(response.context['meta_faturamento'], Decimal('1000.00'))
        self.assertEqual(response.context['percentual_meta_atingido'], Decimal('50.00'))

    def test_dashboard_relatorio_lancamentos_excel(self):
        self._login()

        self._create_lancamento(
            data=date(2026, 3, 10),
            valor_bruto=Decimal('100.00'),
            forma_pagamento=self.forma_debito,
            taxa_percentual=Decimal('3.00'),
        )
        self._create_lancamento(
            data=date(2026, 3, 11),
            valor_bruto=Decimal('80.00'),
            forma_pagamento=self.forma_dinheiro,
            taxa_percentual=Decimal('0.00'),
        )
        self._create_lancamento(
            data=date(2026, 3, 12),
            valor_bruto=Decimal('40.00'),
            permuta=True,
        )
        self._create_lancamento(
            data=date(2026, 2, 5),
            valor_bruto=Decimal('500.00'),
            forma_pagamento=self.forma_pix,
            taxa_percentual=Decimal('0.00'),
        )

        response = self.client.get(
            reverse('salao:dashboard_relatorio_lancamentos'),
            {'ano': 2026, 'mes': 3},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertIn('relatorio_lancamentos_2026_03.xlsx', response['Content-Disposition'])

        workbook = load_workbook(filename=BytesIO(response.content))
        sheet = workbook.active

        headers = [sheet.cell(row=1, column=col).value for col in range(1, 12)]
        self.assertEqual(
            headers,
            [
                'Data',
                'Servico',
                'Tipo',
                'Forma de pagamento',
                'Valor',
                'Parcelas',
                'Taxa (%)',
                'Valor taxa',
                'Valor liquido',
                'Valor 20%',
                'Valor apos 20%',
            ],
        )

        self.assertEqual(sheet.max_row, 6)

        primeira_linha = [sheet.cell(row=2, column=col).value for col in range(1, 12)]
        self.assertEqual(primeira_linha[0], '10/03/2026')
        self.assertEqual(primeira_linha[2], 'NORMAL')
        self.assertEqual(primeira_linha[3], self.forma_debito.nome)
        self.assertAlmostEqual(primeira_linha[4], 100.00, places=2)
        self.assertEqual(primeira_linha[5], 1)
        self.assertAlmostEqual(primeira_linha[6], 3.00, places=2)
        self.assertAlmostEqual(primeira_linha[7], 3.00, places=2)
        self.assertAlmostEqual(primeira_linha[8], 97.00, places=2)
        self.assertAlmostEqual(primeira_linha[9], 19.40, places=2)
        self.assertAlmostEqual(primeira_linha[10], 77.60, places=2)

        linha_permuta = [sheet.cell(row=4, column=col).value for col in range(1, 12)]
        self.assertEqual(linha_permuta[2], 'PERMUTA')
        self.assertAlmostEqual(linha_permuta[4], 40.00, places=2)
        self.assertAlmostEqual(linha_permuta[7], 0.00, places=2)
        self.assertAlmostEqual(linha_permuta[9], 0.00, places=2)
        self.assertAlmostEqual(linha_permuta[10], 40.00, places=2)

        total_linha = [sheet.cell(row=6, column=col).value for col in range(1, 12)]
        self.assertEqual(total_linha[0], 'TOTAL')
        self.assertAlmostEqual(total_linha[4], 220.00, places=2)
        self.assertAlmostEqual(total_linha[7], 3.00, places=2)
        self.assertAlmostEqual(total_linha[8], 217.00, places=2)
        self.assertAlmostEqual(total_linha[9], 35.40, places=2)
        self.assertAlmostEqual(total_linha[10], 181.60, places=2)

    def test_lancamento_taxa_da_cliente_faz_gross_up_do_valor(self):
        """Digita 100 (o que quer receber) e o sistema cobra 103,25 da cliente."""
        self._login()
        TaxaFormaPagamentoSalao.objects.update_or_create(
            forma_pagamento=self.forma_credito, parcelas=1,
            defaults={'percentual': Decimal('3.15')},
        )

        self.client.post(
            reverse('salao:lancamentos'),
            {
                'action': 'create_lancamento',
                'ano': 2026, 'mes': 3, 'dia': 10,
                'codigo': self.servico.codigo,
                'codigo_forma_pagamento': self.forma_credito.codigo,
                'parcelas': 1,
                'valor_bruto': '100,00',
                'taxa_repassada': 'on',
            },
        )

        lancamento = LancamentoSalao.objects.get()
        self.assertTrue(lancamento.taxa_repassada)
        self.assertEqual(lancamento.valor_bruto, Decimal('103.25'))
        self.assertEqual(lancamento.valor_taxa, Decimal('3.25'))
        # o que importa: o salão recebe exatamente os 100 digitados
        self.assertEqual(lancamento.valor_cobrado, Decimal('100.00'))

    def test_lancamento_sem_taxa_da_cliente_mantem_o_comportamento_antigo(self):
        self._login()
        TaxaFormaPagamentoSalao.objects.update_or_create(
            forma_pagamento=self.forma_credito, parcelas=1,
            defaults={'percentual': Decimal('3.15')},
        )

        self.client.post(
            reverse('salao:lancamentos'),
            {
                'action': 'create_lancamento',
                'ano': 2026, 'mes': 3, 'dia': 10,
                'codigo': self.servico.codigo,
                'codigo_forma_pagamento': self.forma_credito.codigo,
                'parcelas': 1,
                'valor_bruto': '100,00',
            },
        )

        lancamento = LancamentoSalao.objects.get()
        self.assertFalse(lancamento.taxa_repassada)
        self.assertEqual(lancamento.valor_bruto, Decimal('100.00'))
        self.assertEqual(lancamento.valor_cobrado, Decimal('96.85'))

    def test_lancamento_taxa_da_cliente_ignorada_na_permuta(self):
        self._login()

        self.client.post(
            reverse('salao:lancamentos'),
            {
                'action': 'create_lancamento',
                'ano': 2026, 'mes': 3, 'dia': 10,
                'codigo': self.servico.codigo,
                'parcelas': 1,
                'valor_bruto': '100,00',
                'permuta': 'on',
                'taxa_repassada': 'on',
            },
        )

        lancamento = LancamentoSalao.objects.get()
        self.assertTrue(lancamento.permuta)
        self.assertFalse(lancamento.taxa_repassada)
        self.assertEqual(lancamento.valor_bruto, Decimal('100.00'))
        self.assertEqual(lancamento.valor_cobrado, Decimal('100.00'))

    def test_gross_up_entrega_sempre_o_liquido_pedido(self):
        for alvo, percentual in [
            ('100.00', '3.15'), ('160.00', '7.19'), ('80.00', '1.37'),
            ('200.00', '5.39'), ('1234.56', '12.40'), ('100.00', '0.00'),
        ]:
            with self.subTest(alvo=alvo, taxa=percentual):
                bruto = _calcular_bruto_com_taxa_repassada(Decimal(alvo), Decimal(percentual))
                _, liquido = _calcular_liquido_com_taxa(bruto, Decimal(percentual))
                self.assertEqual(liquido, Decimal(alvo))

    def test_editar_lancamento_volta_para_o_dia_do_lancamento(self):
        """Editar um lançamento do dia 15 não pode cair no dia de hoje."""
        self._login()
        lancamento = self._create_lancamento(
            data=date(2026, 3, 15),
            valor_bruto=Decimal('100.00'),
            forma_pagamento=self.forma_dinheiro,
        )

        response = self.client.post(
            reverse('salao:lancamentos'),
            {
                'action': 'update_lancamento',
                'ano': 2026, 'mes': 3,
                'dia': 1,  # a tela estava em outro dia
                'lancamento_id': lancamento.id,
                'data': '2026-03-15',
                'servico_id': self.servico.id,
                'valor_bruto': '120,00',
                'forma_pagamento_id': self.forma_dinheiro.id,
                'codigo_forma_pagamento': self.forma_dinheiro.codigo,
                'parcelas': 1,
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertIn('dia=15', response['Location'])
        lancamento.refresh_from_db()
        self.assertEqual(lancamento.valor_bruto, Decimal('120.00'))

    def test_editar_lancamento_mudando_a_data_segue_o_lancamento(self):
        self._login()
        lancamento = self._create_lancamento(
            data=date(2026, 3, 15),
            valor_bruto=Decimal('100.00'),
            forma_pagamento=self.forma_dinheiro,
        )

        response = self.client.post(
            reverse('salao:lancamentos'),
            {
                'action': 'update_lancamento',
                'ano': 2026, 'mes': 3, 'dia': 1,
                'lancamento_id': lancamento.id,
                'data': '2026-04-20',
                'servico_id': self.servico.id,
                'valor_bruto': '100,00',
                'forma_pagamento_id': self.forma_dinheiro.id,
                'codigo_forma_pagamento': self.forma_dinheiro.codigo,
                'parcelas': 1,
            },
        )
        self.assertIn('ano=2026', response['Location'])
        self.assertIn('mes=4', response['Location'])
        self.assertIn('dia=20', response['Location'])

    def test_editar_lancamento_com_erro_volta_para_o_dia_de_origem(self):
        self._login()
        lancamento = self._create_lancamento(
            data=date(2026, 3, 15),
            valor_bruto=Decimal('100.00'),
            forma_pagamento=self.forma_dinheiro,
        )

        response = self.client.post(
            reverse('salao:lancamentos'),
            {
                'action': 'update_lancamento',
                'ano': 2026, 'mes': 3, 'dia': 1,
                'lancamento_id': lancamento.id,
                'data': 'data-invalida',
                'servico_id': self.servico.id,
                'valor_bruto': '100,00',
                'codigo_forma_pagamento': self.forma_dinheiro.codigo,
                'parcelas': 1,
            },
        )
        self.assertIn('dia=15', response['Location'])

    def test_excluir_lancamento_volta_para_o_dia_do_lancamento(self):
        self._login()
        lancamento = self._create_lancamento(
            data=date(2026, 3, 15),
            valor_bruto=Decimal('100.00'),
            forma_pagamento=self.forma_dinheiro,
        )

        response = self.client.post(
            reverse('salao:lancamentos'),
            {
                'action': 'delete_lancamento',
                'ano': 2026, 'mes': 3, 'dia': 1,
                'lancamento_id': lancamento.id,
            },
        )
        self.assertIn('dia=15', response['Location'])

    def test_lancamento_codigo_invalido_bloqueia_save(self):
        self._login()

        response = self.client.post(
            reverse('salao:lancamentos'),
            {
                'action': 'create_lancamento',
                'ano': 2026,
                'mes': 3,
                'dia': 10,
                'codigo': 'X99',
                'codigo_forma_pagamento': '3',
                'parcelas': 1,
                'valor_bruto': '70,00',
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(LancamentoSalao.objects.count(), 0)

    def test_lancamento_forma_invalida_bloqueia_save(self):
        self._login()

        response = self.client.post(
            reverse('salao:lancamentos'),
            {
                'action': 'create_lancamento',
                'ano': 2026,
                'mes': 3,
                'dia': 10,
                'codigo': 'C01',
                'codigo_forma_pagamento': '999',
                'parcelas': 1,
                'valor_bruto': '70,00',
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(LancamentoSalao.objects.count(), 0)

    def test_lancamento_taxa_ausente_bloqueia_save(self):
        self._login()

        response = self.client.post(
            reverse('salao:lancamentos'),
            {
                'action': 'create_lancamento',
                'ano': 2026,
                'mes': 3,
                'dia': 10,
                'codigo': 'C01',
                'codigo_forma_pagamento': '2',
                'parcelas': 12,
                'valor_bruto': '100,00',
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(LancamentoSalao.objects.count(), 0)

    def test_lancamento_codigo_valido_salva_com_taxa(self):
        self._login()

        response = self.client.post(
            reverse('salao:lancamentos'),
            {
                'action': 'create_lancamento',
                'ano': 2026,
                'mes': 3,
                'dia': 10,
                'codigo': 'c01',
                'codigo_forma_pagamento': '4',
                'parcelas': 1,
                'valor_bruto': '100,00',
            },
        )

        self.assertEqual(response.status_code, 302)
        lanc = LancamentoSalao.objects.get()
        self.assertEqual(lanc.valor_bruto, Decimal('100.00'))
        self.assertEqual(lanc.valor_taxa, Decimal('3.00'))
        self.assertEqual(lanc.valor_cobrado, Decimal('97.00'))
        self.assertFalse(lanc.permuta)

    def test_lancamento_permuta_forca_nao_informado_sem_taxa(self):
        self._login()

        response = self.client.post(
            reverse('salao:lancamentos'),
            {
                'action': 'create_lancamento',
                'ano': 2026,
                'mes': 3,
                'dia': 10,
                'codigo': 'c01',
                'codigo_forma_pagamento': '4',
                'parcelas': 12,
                'permuta': 'on',
                'valor_bruto': '120,00',
            },
        )

        self.assertEqual(response.status_code, 302)
        lanc = LancamentoSalao.objects.get()
        self.assertTrue(lanc.permuta)
        self.assertEqual(lanc.forma_pagamento_id, self.forma_nao_informado.id)
        self.assertEqual(lanc.parcelas, 1)
        self.assertEqual(lanc.taxa_percentual_aplicada, Decimal('0.00'))
        self.assertEqual(lanc.valor_taxa, Decimal('0.00'))
        self.assertEqual(lanc.valor_cobrado, Decimal('120.00'))

    def test_endpoint_refresh_lancamentos_por_dia_retorna_rows_html(self):
        self._login()

        self._create_lancamento(
            data=date(2026, 3, 10),
            valor_bruto=Decimal('80.00'),
            forma_pagamento=self.forma_dinheiro,
            taxa_percentual=Decimal('0.00'),
        )

        response = self.client.get(
            reverse('salao:lancamentos'),
            {
                'ano': 2026,
                'mes': 3,
                'dia': 10,
                'refresh': 1,
            },
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertIn('rows_html', payload)
        self.assertIn('Corte Feminino', payload['rows_html'])

    def test_editar_e_excluir_lancamento(self):
        self._login()

        lancamento = self._create_lancamento(
            data=date(2026, 3, 10),
            valor_bruto=Decimal('80.00'),
            forma_pagamento=self.forma_dinheiro,
            taxa_percentual=Decimal('0.00'),
        )

        edit_response = self.client.post(
            reverse('salao:lancamentos'),
            {
                'action': 'update_lancamento',
                'lancamento_id': lancamento.id,
                'ano': 2026,
                'mes': 3,
                'dia': 10,
                'data': '2026-03-11',
                'servico_id': self.servico.id,
                'forma_pagamento_id': self.forma_debito.id,
                'parcelas': 1,
                'valor_bruto': '95,00',
            },
        )
        self.assertEqual(edit_response.status_code, 302)

        lancamento.refresh_from_db()
        self.assertEqual(lancamento.data, date(2026, 3, 11))
        self.assertEqual(lancamento.valor_bruto, Decimal('95.00'))
        self.assertEqual(lancamento.valor_taxa, Decimal('2.85'))
        self.assertEqual(lancamento.valor_cobrado, Decimal('92.15'))

        delete_response = self.client.post(
            reverse('salao:lancamentos'),
            {
                'action': 'delete_lancamento',
                'lancamento_id': lancamento.id,
                'ano': 2026,
                'mes': 3,
                'dia': 10,
            },
        )
        self.assertEqual(delete_response.status_code, 302)
        self.assertEqual(LancamentoSalao.objects.count(), 0)

    def test_editar_lancamento_alterna_normal_e_permuta(self):
        self._login()
        lancamento = self._create_lancamento(
            data=date(2026, 3, 10),
            valor_bruto=Decimal('120.00'),
            forma_pagamento=self.forma_credito,
            parcelas=2,
            taxa_percentual=Decimal('5.00'),
        )

        to_permuta = self.client.post(
            reverse('salao:lancamentos'),
            {
                'action': 'update_lancamento',
                'lancamento_id': lancamento.id,
                'ano': 2026,
                'mes': 3,
                'dia': 10,
                'data': '2026-03-10',
                'servico_id': self.servico.id,
                'forma_pagamento_id': self.forma_credito.id,
                'parcelas': 2,
                'permuta': 'on',
                'valor_bruto': '200,00',
            },
        )
        self.assertEqual(to_permuta.status_code, 302)

        lancamento.refresh_from_db()
        self.assertTrue(lancamento.permuta)
        self.assertEqual(lancamento.forma_pagamento_id, self.forma_nao_informado.id)
        self.assertEqual(lancamento.parcelas, 1)
        self.assertEqual(lancamento.valor_taxa, Decimal('0.00'))
        self.assertEqual(lancamento.valor_cobrado, Decimal('200.00'))

        to_normal = self.client.post(
            reverse('salao:lancamentos'),
            {
                'action': 'update_lancamento',
                'lancamento_id': lancamento.id,
                'ano': 2026,
                'mes': 3,
                'dia': 10,
                'data': '2026-03-11',
                'servico_id': self.servico.id,
                'forma_pagamento_id': self.forma_debito.id,
                'parcelas': 1,
                'valor_bruto': '200,00',
            },
        )
        self.assertEqual(to_normal.status_code, 302)
        lancamento.refresh_from_db()
        self.assertFalse(lancamento.permuta)
        self.assertEqual(lancamento.forma_pagamento_id, self.forma_debito.id)
        self.assertEqual(lancamento.parcelas, 1)
        self.assertEqual(lancamento.taxa_percentual_aplicada, Decimal('3.00'))
        self.assertEqual(lancamento.valor_taxa, Decimal('6.00'))
        self.assertEqual(lancamento.valor_cobrado, Decimal('194.00'))

    def test_criar_despesa_parcelada_em_4x(self):
        self._login()

        response = self.client.post(
            reverse('salao:despesas'),
            {
                'action': 'create_despesa',
                'ano': 2026,
                'mes': 3,
                'data': '2026-03-31',
                'categoria_id': self.categoria.id,
                'valor': '2000,00',
                'parcelas': 4,
                'observacao': 'Cadeira nova',
            },
        )
        self.assertEqual(response.status_code, 302)

        despesas = list(DespesaSalao.objects.order_by('data', 'parcela_numero'))
        self.assertEqual(len(despesas), 4)
        self.assertEqual([d.parcela_numero for d in despesas], [1, 2, 3, 4])
        self.assertEqual([d.parcelas_total for d in despesas], [4, 4, 4, 4])
        self.assertEqual(despesas[0].data, date(2026, 3, 31))
        self.assertEqual(despesas[1].data, date(2026, 4, 30))
        self.assertEqual(despesas[2].data, date(2026, 5, 31))
        self.assertEqual(despesas[3].data, date(2026, 6, 30))
        self.assertEqual(sum(d.valor for d in despesas), Decimal('2000.00'))
        self.assertIsNotNone(despesas[0].grupo_parcelamento_id)

    def test_excluir_grupo_despesa_parcelada(self):
        self._login()

        self.client.post(
            reverse('salao:despesas'),
            {
                'action': 'create_despesa',
                'ano': 2026,
                'mes': 3,
                'data': '2026-03-10',
                'categoria_id': self.categoria.id,
                'valor': '300,00',
                'parcelas': 3,
                'observacao': 'Teste',
            },
        )
        first = DespesaSalao.objects.first()
        self.assertIsNotNone(first.grupo_parcelamento_id)

        response = self.client.post(
            reverse('salao:despesas'),
            {
                'action': 'delete_despesa_grupo',
                'grupo_parcelamento_id': str(first.grupo_parcelamento_id),
                'ano': 2026,
                'mes': 3,
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(DespesaSalao.objects.count(), 0)

    def test_criar_editar_excluir_despesa(self):
        self._login()

        create_response = self.client.post(
            reverse('salao:despesas'),
            {
                'action': 'create_despesa',
                'ano': 2026,
                'mes': 3,
                'data': '2026-03-14',
                'categoria_id': self.categoria.id,
                'valor': '45,00',
                'parcelas': 1,
                'observacao': 'Luva descartável',
            },
        )
        self.assertEqual(create_response.status_code, 302)
        self.assertEqual(DespesaSalao.objects.count(), 1)

        despesa = DespesaSalao.objects.first()

        edit_response = self.client.post(
            reverse('salao:despesas'),
            {
                'action': 'update_despesa',
                'despesa_id': despesa.id,
                'ano': 2026,
                'mes': 3,
                'data': '2026-03-15',
                'categoria_id': self.categoria.id,
                'valor': '55,00',
                'observacao': 'Produto extra',
            },
        )
        self.assertEqual(edit_response.status_code, 302)

        despesa.refresh_from_db()
        self.assertEqual(despesa.data, date(2026, 3, 15))
        self.assertEqual(despesa.valor, Decimal('55.00'))

        delete_response = self.client.post(
            reverse('salao:despesas'),
            {
                'action': 'delete_despesa',
                'despesa_id': despesa.id,
                'ano': 2026,
                'mes': 3,
            },
        )
        self.assertEqual(delete_response.status_code, 302)
        self.assertEqual(DespesaSalao.objects.count(), 0)

    def test_crud_servicos_pela_tela(self):
        self._login()

        create_response = self.client.post(
            reverse('salao:servicos'),
            {
                'action': 'create_servico',
                'codigo': 'P10',
                'nome': 'Progressiva',
                'valor_padrao': '250,00',
                'ativo': 'on',
            },
        )
        self.assertEqual(create_response.status_code, 302)
        created = ServicoSalao.objects.get(codigo='P10')
        self.assertEqual(created.valor_padrao, Decimal('250.00'))

        update_response = self.client.post(
            reverse('salao:servicos'),
            {
                'action': 'update_servico',
                'servico_id': created.id,
                'codigo': 'P10',
                'nome': 'Progressiva Premium',
                'valor_padrao': '300,00',
            },
        )
        self.assertEqual(update_response.status_code, 302)
        created.refresh_from_db()
        self.assertEqual(created.nome, 'Progressiva Premium')
        self.assertEqual(created.valor_padrao, Decimal('300.00'))
        self.assertFalse(created.ativo)

    def test_crud_categorias_pela_tela(self):
        self._login()

        create_response = self.client.post(
            reverse('salao:categorias'),
            {
                'action': 'create_categoria',
                'nome': 'Lavanderia',
                'ativo': 'on',
            },
        )
        self.assertEqual(create_response.status_code, 302)
        created = CategoriaDespesaSalao.objects.get(nome='Lavanderia')
        self.assertTrue(created.ativo)

        update_response = self.client.post(
            reverse('salao:categorias'),
            {
                'action': 'update_categoria',
                'categoria_id': created.id,
                'nome': 'Lavanderia e Toalhas',
            },
        )
        self.assertEqual(update_response.status_code, 302)
        created.refresh_from_db()
        self.assertEqual(created.nome, 'Lavanderia e Toalhas')
        self.assertFalse(created.ativo)

    def test_crud_subcategorias_pela_tela(self):
        self._login()
        create_response = self.client.post(
            reverse('salao:categorias'),
            {
                'action': 'create_subcategoria',
                'categoria_id': self.categoria.id,
                'nome': 'Auxiliar Joana',
                'ativo': 'on',
            },
        )
        self.assertEqual(create_response.status_code, 302)
        created = SubcategoriaDespesaSalao.objects.get(nome='Auxiliar Joana')
        self.assertTrue(created.ativo)

        update_response = self.client.post(
            reverse('salao:categorias'),
            {
                'action': 'update_subcategoria',
                'subcategoria_id': created.id,
                'categoria_id': self.categoria.id,
                'nome': 'Auxiliar Joana Silva',
            },
        )
        self.assertEqual(update_response.status_code, 302)
        created.refresh_from_db()
        self.assertEqual(created.nome, 'Auxiliar Joana Silva')
        self.assertFalse(created.ativo)

    def test_crud_pagamentos_e_taxas(self):
        self._login()

        create_response = self.client.post(
            reverse('salao:pagamentos'),
            {
                'action': 'create_forma_pagamento',
                'codigo': '9',
                'nome': 'Carteira Digital',
                'aceita_parcelamento': 'on',
                'ativo': 'on',
            },
        )
        self.assertEqual(create_response.status_code, 302)
        forma = FormaPagamentoSalao.objects.get(codigo='9')

        save_taxas_response = self.client.post(
            reverse('salao:pagamentos'),
            {
                'action': 'save_taxas_forma',
                'forma_id': forma.id,
                'taxa_1': '1,5',
                'taxa_2': '2,5',
            },
        )
        self.assertEqual(save_taxas_response.status_code, 302)
        self.assertTrue(TaxaFormaPagamentoSalao.objects.filter(forma_pagamento=forma, parcelas=1).exists())
        self.assertTrue(TaxaFormaPagamentoSalao.objects.filter(forma_pagamento=forma, parcelas=2).exists())

    def test_despesa_normal_nao_movimenta_estoque(self):
        self._login()

        response = self.client.post(
            reverse('salao:despesas'),
            {
                'action': 'create_despesa',
                'ano': 2026,
                'mes': 3,
                'data': '2026-03-10',
                'categoria_id': self.categoria.id,
                'valor': '120,00',
                'parcelas': 1,
                'observacao': 'Energia elétrica',
            },
        )
        self.assertEqual(response.status_code, 302)
        despesa = DespesaSalao.objects.get()
        self.assertFalse(despesa.gera_estoque)
        self.assertEqual(MovimentoEstoqueSalao.objects.count(), 0)
        self.produto.refresh_from_db()
        self.assertEqual(self.produto.saldo_atual, Decimal('0.000'))

    def test_despesa_salva_subcategoria_quando_informada(self):
        self._login()
        response = self.client.post(
            reverse('salao:despesas'),
            {
                'action': 'create_despesa',
                'ano': 2026,
                'mes': 3,
                'data': '2026-03-21',
                'categoria_id': self.categoria.id,
                'subcategoria_id': self.subcategoria.id,
                'valor': '130,00',
                'parcelas': 1,
                'observacao': 'Pagamento auxiliar',
            },
        )
        self.assertEqual(response.status_code, 302)
        despesa = DespesaSalao.objects.get()
        self.assertEqual(despesa.subcategoria_id, self.subcategoria.id)

    def test_despesa_com_estoque_cria_compra_itens_e_movimento(self):
        self._login()

        response = self.client.post(
            reverse('salao:despesas'),
            {
                'action': 'create_despesa',
                'ano': 2026,
                'mes': 3,
                'data': '2026-03-10',
                'categoria_id': self.categoria.id,
                'parcelas': 1,
                'observacao': 'Compra fornecedor RIGOLIM',
                'gera_estoque': 'on',
                'produto_id[]': [str(self.produto.id)],
                'quantidade[]': ['10'],
                'custo_unitario[]': ['20,00'],
            },
        )
        self.assertEqual(response.status_code, 302)

        despesa = DespesaSalao.objects.get()
        self.assertTrue(despesa.gera_estoque)
        self.assertEqual(despesa.valor, Decimal('200.00'))
        self.assertEqual(CompraEstoqueSalao.objects.count(), 1)
        self.assertEqual(CompraEstoqueItemSalao.objects.count(), 1)
        self.assertEqual(MovimentoEstoqueSalao.objects.count(), 1)

        item = CompraEstoqueItemSalao.objects.first()
        self.assertEqual(item.quantidade, Decimal('10.000'))
        self.assertEqual(item.custo_total, Decimal('200.00'))

        mov = MovimentoEstoqueSalao.objects.first()
        self.assertEqual(mov.tipo, MovimentoEstoqueSalao.TIPO_ENTRADA)
        self.assertEqual(mov.motivo, MovimentoEstoqueSalao.MOTIVO_COMPRA)
        self.assertEqual(mov.valor_custo_total, Decimal('200.00'))

        self.produto.refresh_from_db()
        self.assertEqual(self.produto.saldo_atual, Decimal('10.000'))
        self.assertEqual(self.produto.custo_medio_atual, Decimal('20.00'))

    def test_compra_parcelada_gera_entrada_fisica_unica(self):
        self._login()

        response = self.client.post(
            reverse('salao:despesas'),
            {
                'action': 'create_despesa',
                'ano': 2026,
                'mes': 3,
                'data': '2026-03-05',
                'categoria_id': self.categoria.id,
                'parcelas': 4,
                'observacao': 'Compra parcelada',
                'gera_estoque': 'on',
                'produto_id[]': [str(self.produto.id)],
                'quantidade[]': ['8'],
                'custo_unitario[]': ['10,00'],
            },
        )
        self.assertEqual(response.status_code, 302)

        despesas = DespesaSalao.objects.order_by('parcela_numero')
        self.assertEqual(despesas.count(), 4)
        self.assertEqual(sum(d.valor for d in despesas), Decimal('80.00'))
        self.assertTrue(all(d.gera_estoque for d in despesas))
        self.assertEqual(MovimentoEstoqueSalao.objects.filter(motivo=MovimentoEstoqueSalao.MOTIVO_COMPRA).count(), 1)
        self.produto.refresh_from_db()
        self.assertEqual(self.produto.saldo_atual, Decimal('8.000'))

    def test_saida_venda_com_taxa_calcula_liquido_e_lucro(self):
        self._login()
        self.produto.saldo_atual = Decimal('5.000')
        self.produto.custo_medio_atual = Decimal('10.00')
        self.produto.save(update_fields=['saldo_atual', 'custo_medio_atual', 'atualizado_em'])

        response = self.client.post(
            reverse('salao:estoque'),
            {
                'action': 'create_saida_estoque',
                'ano': 2026,
                'mes': 3,
                'data': '2026-03-12',
                'produto_id': self.produto.id,
                'tipo_saida': 'VENDA',
                'quantidade': '2',
                'valor_venda_unitario': '30,00',
                'forma_pagamento_id': self.forma_debito.id,
                'parcelas': 1,
            },
        )
        self.assertEqual(response.status_code, 302)

        mov = MovimentoEstoqueSalao.objects.get(motivo=MovimentoEstoqueSalao.MOTIVO_VENDA)
        self.assertEqual(mov.valor_bruto_venda, Decimal('60.00'))
        self.assertEqual(mov.taxa_percentual_aplicada, Decimal('3.00'))
        self.assertEqual(mov.valor_taxa, Decimal('1.80'))
        self.assertEqual(mov.valor_liquido_venda, Decimal('58.20'))
        self.assertEqual(mov.valor_custo_total, Decimal('20.00'))
        self.assertEqual(mov.lucro_produto, Decimal('38.20'))

        self.produto.refresh_from_db()
        self.assertEqual(self.produto.saldo_atual, Decimal('3.000'))

    def test_saida_venda_sem_taxa_cadastrada_aplica_zero(self):
        self._login()
        self.produto.saldo_atual = Decimal('4.000')
        self.produto.custo_medio_atual = Decimal('10.00')
        self.produto.save(update_fields=['saldo_atual', 'custo_medio_atual', 'atualizado_em'])

        response = self.client.post(
            reverse('salao:estoque'),
            {
                'action': 'create_saida_estoque',
                'ano': 2026,
                'mes': 3,
                'data': '2026-03-12',
                'produto_id': self.produto.id,
                'tipo_saida': 'VENDA',
                'quantidade': '1',
                'valor_venda_unitario': '40,00',
                'forma_pagamento_id': self.forma_credito.id,
                'parcelas': 12,
            },
        )
        self.assertEqual(response.status_code, 302)

        mov = MovimentoEstoqueSalao.objects.get(motivo=MovimentoEstoqueSalao.MOTIVO_VENDA)
        self.assertEqual(mov.taxa_percentual_aplicada, Decimal('0.00'))
        self.assertEqual(mov.valor_taxa, Decimal('0.00'))
        self.assertEqual(mov.valor_liquido_venda, Decimal('40.00'))

    def test_saida_uso_em_cliente_baixa_sem_receita(self):
        self._login()
        self.produto.saldo_atual = Decimal('4.000')
        self.produto.custo_medio_atual = Decimal('12.00')
        self.produto.save(update_fields=['saldo_atual', 'custo_medio_atual', 'atualizado_em'])

        response = self.client.post(
            reverse('salao:estoque'),
            {
                'action': 'create_saida_estoque',
                'ano': 2026,
                'mes': 3,
                'data': '2026-03-15',
                'produto_id': self.produto.id,
                'tipo_saida': 'USO_EM_CLIENTE',
                'quantidade': '1,5',
            },
        )
        self.assertEqual(response.status_code, 302)

        mov = MovimentoEstoqueSalao.objects.get(motivo=MovimentoEstoqueSalao.MOTIVO_USO_EM_CLIENTE)
        self.assertEqual(mov.valor_liquido_venda, Decimal('0.00'))
        self.assertEqual(mov.valor_custo_total, Decimal('18.00'))
        self.produto.refresh_from_db()
        self.assertEqual(self.produto.saldo_atual, Decimal('2.500'))

    def test_saida_bloqueia_estoque_insuficiente(self):
        self._login()
        self.produto.saldo_atual = Decimal('1.000')
        self.produto.custo_medio_atual = Decimal('10.00')
        self.produto.save(update_fields=['saldo_atual', 'custo_medio_atual', 'atualizado_em'])

        response = self.client.post(
            reverse('salao:estoque'),
            {
                'action': 'create_saida_estoque',
                'ano': 2026,
                'mes': 3,
                'data': '2026-03-18',
                'produto_id': self.produto.id,
                'tipo_saida': 'USO_EM_CLIENTE',
                'quantidade': '2',
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(MovimentoEstoqueSalao.objects.count(), 0)
        self.produto.refresh_from_db()
        self.assertEqual(self.produto.saldo_atual, Decimal('1.000'))

    def test_dashboard_produto_separado_sem_comissao_de_produto(self):
        self._login()
        self._create_lancamento(
            data=date(2026, 3, 10),
            valor_bruto=Decimal('100.00'),
            forma_pagamento=self.forma_dinheiro,
            taxa_percentual=Decimal('0.00'),
        )
        MovimentoEstoqueSalao.objects.create(
            data=date(2026, 3, 11),
            produto=self.produto,
            tipo=MovimentoEstoqueSalao.TIPO_SAIDA,
            motivo=MovimentoEstoqueSalao.MOTIVO_VENDA,
            quantidade=Decimal('1.000'),
            custo_unitario_aplicado=Decimal('10.00'),
            valor_custo_total=Decimal('10.00'),
            valor_venda_unitario=Decimal('50.00'),
            valor_bruto_venda=Decimal('50.00'),
            taxa_percentual_aplicada=Decimal('0.00'),
            valor_taxa=Decimal('0.00'),
            valor_liquido_venda=Decimal('50.00'),
            lucro_produto=Decimal('40.00'),
            forma_pagamento=self.forma_dinheiro,
            parcelas=1,
        )

        response = self.client.get(reverse('salao:dashboard'), {'ano': 2026, 'mes': 3})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['comissao_calculada'], Decimal('20.00'))
        self.assertEqual(response.context['vendas_produto_liquidas'], Decimal('50.00'))
        self.assertEqual(response.context['lucro_produto'], Decimal('40.00'))

    def test_dashboard_permuta_separada_sem_impactar_lucro(self):
        self._login()
        self._create_lancamento(
            data=date(2026, 3, 10),
            valor_bruto=Decimal('100.00'),
            forma_pagamento=self.forma_dinheiro,
            taxa_percentual=Decimal('0.00'),
        )
        self._create_lancamento(
            data=date(2026, 3, 11),
            valor_bruto=Decimal('80.00'),
            permuta=True,
        )
        DespesaSalao.objects.create(
            data=date(2026, 3, 12),
            categoria=self.categoria,
            valor=Decimal('30.00'),
        )

        response = self.client.get(reverse('salao:dashboard'), {'ano': 2026, 'mes': 3})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['faturamento_bruto'], Decimal('100.00'))
        self.assertEqual(response.context['permuta_total_mes'], Decimal('80.00'))
        self.assertEqual(response.context['comissao_calculada'], Decimal('20.00'))
        self.assertEqual(response.context['lucro'], Decimal('50.00'))
        self.assertEqual(response.context['atendimentos_total'], 2)
        self.assertEqual(response.context['ticket_permuta'], Decimal('80.00'))
        self.assertEqual(response.context['atendimentos_permuta_total'], 1)
        self.assertIn('permuta_chart', response.context)

    def test_dashboard_sem_comissao_nao_entra_na_base_da_comissao(self):
        self._login()
        self._create_lancamento(
            data=date(2026, 3, 10),
            valor_bruto=Decimal('100.00'),
            forma_pagamento=self.forma_dinheiro,
            taxa_percentual=Decimal('0.00'),
        )
        self._create_lancamento(
            data=date(2026, 3, 11),
            valor_bruto=Decimal('50.00'),
            forma_pagamento=self.forma_dinheiro,
            taxa_percentual=Decimal('0.00'),
            sem_comissao=True,
        )

        response = self.client.get(reverse('salao:dashboard'), {'ano': 2026, 'mes': 3})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['faturamento_bruto'], Decimal('100.00'))
        self.assertEqual(response.context['sem_comissao_total_mes'], Decimal('50.00'))
        self.assertEqual(response.context['atendimentos_sem_comissao_total'], 1)
        self.assertEqual(response.context['ticket_sem_comissao'], Decimal('50.00'))
        self.assertEqual(response.context['comissao_calculada'], Decimal('20.00'))
        self.assertEqual(response.context['valor_pos_comissao'], Decimal('80.00'))

    def test_dashboard_agrupa_despesas_por_subcategoria(self):
        self._login()
        DespesaSalao.objects.create(
            data=date(2026, 3, 10),
            categoria=self.categoria,
            subcategoria=self.subcategoria,
            valor=Decimal('90.00'),
            observacao='Auxiliar',
        )
        response = self.client.get(reverse('salao:dashboard'), {'ano': 2026, 'mes': 3})
        self.assertEqual(response.status_code, 200)
        rows = list(response.context['despesas_por_subcategoria'])
        self.assertTrue(any(item['subcategoria_nome'] == self.subcategoria.nome for item in rows))

    def test_grid_lancamentos_filtra_por_servico_e_pagamento(self):
        self._login()
        outro_servico = ServicoSalao.objects.create(
            codigo='C99',
            nome='Outro Serviço',
            valor_padrao=Decimal('40.00'),
            ativo=True,
        )
        self._create_lancamento(
            data=date(2026, 3, 10),
            valor_bruto=Decimal('80.00'),
            forma_pagamento=self.forma_dinheiro,
            taxa_percentual=Decimal('0.00'),
        )
        LancamentoSalao.objects.create(
            data=date(2026, 3, 11),
            servico=outro_servico,
            forma_pagamento=self.forma_credito,
            parcelas=1,
            valor_bruto=Decimal('90.00'),
            taxa_percentual_aplicada=Decimal('4.00'),
            valor_taxa=Decimal('3.60'),
            valor_cobrado=Decimal('86.40'),
        )
        response = self.client.get(
            reverse('salao:grid_lancamentos'),
            {'ano': 2026, 'mes': 3, 'servico_id': self.servico.id, 'forma_pagamento_id': self.forma_dinheiro.id},
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.context['page_obj'].object_list), 1)

    def test_grid_lancamentos_exibe_permuta(self):
        self._login()
        self._create_lancamento(
            data=date(2026, 3, 10),
            valor_bruto=Decimal('90.00'),
            permuta=True,
        )
        response = self.client.get(
            reverse('salao:grid_lancamentos'),
            {'ano': 2026, 'mes': 3},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'PERMUTA')

    def test_grid_despesas_filtra_por_categoria_e_subcategoria(self):
        self._login()
        outra_categoria = CategoriaDespesaSalao.objects.create(nome='Fixos', ativo=True)
        DespesaSalao.objects.create(
            data=date(2026, 3, 9),
            categoria=self.categoria,
            subcategoria=self.subcategoria,
            valor=Decimal('70.00'),
        )
        DespesaSalao.objects.create(
            data=date(2026, 3, 9),
            categoria=outra_categoria,
            valor=Decimal('30.00'),
        )
        response = self.client.get(
            reverse('salao:grid_despesas'),
            {'ano': 2026, 'mes': 3, 'categoria_id': self.categoria.id, 'subcategoria_id': self.subcategoria.id},
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.context['page_obj'].object_list), 1)

    def test_alerta_estoque_minimo_em_estoque_e_dashboard(self):
        self._login()
        self.produto.saldo_atual = Decimal('1.000')
        self.produto.estoque_minimo = Decimal('2.000')
        self.produto.save(update_fields=['saldo_atual', 'estoque_minimo', 'atualizado_em'])

        estoque_response = self.client.get(reverse('salao:estoque'), {'ano': 2026, 'mes': 3})
        self.assertEqual(estoque_response.status_code, 200)
        self.assertIn(self.produto, list(estoque_response.context['produtos_alerta']))

        dashboard_response = self.client.get(reverse('salao:dashboard'), {'ano': 2026, 'mes': 3})
        self.assertEqual(dashboard_response.status_code, 200)
        self.assertIn(self.produto, list(dashboard_response.context['produtos_alerta']))
